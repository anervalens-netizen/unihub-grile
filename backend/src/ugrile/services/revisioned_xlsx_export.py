"""Revision-pinned XLSX rendering for durable export jobs.

The original S5 renderer predates historical Pontaj/grid revisions and can
select rows from more than one revision. Durable jobs must never mix those
snapshots. These renderers take an explicit calendar-derived ``revision`` and
apply it to every revisioned read used to build the artifact.

This module deliberately reuses the established workbook-formatting helpers;
it changes selection semantics, not the approved workbook layout.
"""

from __future__ import annotations

import json
from collections.abc import Iterable
from datetime import date
from decimal import Decimal
from types import SimpleNamespace
from typing import cast

from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from ..repositories.models import (
    GridCalculation,
    Month,
    Person,
    PontajProjection,
    SalesPersonDayProjection,
    SiteDayAssignment,
    Store,
)
from .xlsx_artifacts import deterministic_zip, save_workbook_deterministic
from .xlsx_export import (
    SCHEMA,
    ExportEnvelope,
    _checksum,
    _determine_generation,
    _safe_filename,
    _write_grila_tab,
    _write_pontaj_tab,
)


def _assignments(
    session: Session,
    *,
    tenant_id: str,
    month_id: str,
    revision: int,
    store_ids: Iterable[str] | None = None,
) -> list[SiteDayAssignment]:
    query = select(SiteDayAssignment).where(
        SiteDayAssignment.tenant_id == tenant_id,
        SiteDayAssignment.month_id == month_id,
        SiteDayAssignment.revision == revision,
    )
    if store_ids is not None:
        query = query.where(SiteDayAssignment.store_id.in_(list(store_ids)))
    return list(
        session.execute(
            query.order_by(SiteDayAssignment.person_id, SiteDayAssignment.business_date)
        ).scalars()
    )


def _pontaj(
    session: Session,
    *,
    tenant_id: str,
    month_id: str,
    revision: int,
    person_ids: set[str],
) -> list[PontajProjection]:
    if not person_ids:
        return []
    return list(
        session.execute(
            select(PontajProjection).where(
                PontajProjection.tenant_id == tenant_id,
                PontajProjection.month_id == month_id,
                PontajProjection.revision == revision,
                PontajProjection.person_id.in_(person_ids),
            )
        ).scalars()
    )


def _people(
    session: Session,
    *,
    tenant_id: str,
    person_ids: set[str],
) -> list[Person]:
    if not person_ids:
        return []
    rows = list(
        session.execute(
            select(Person).where(
                Person.tenant_id == tenant_id,
                Person.id.in_(person_ids),
            )
        ).scalars()
    )
    rows.sort(key=lambda person: ((person.internal_code or ""), person.id))
    return rows


def _sales(
    session: Session,
    *,
    tenant_id: str,
    month_id: str,
    store_id: str,
    revision: int,
) -> dict[tuple[str, date], Decimal]:
    rows = list(
        session.execute(
            select(SalesPersonDayProjection).where(
                SalesPersonDayProjection.tenant_id == tenant_id,
                SalesPersonDayProjection.month_id == month_id,
                SalesPersonDayProjection.store_id == store_id,
                SalesPersonDayProjection.revision == revision,
            )
        ).scalars()
    )
    result: dict[tuple[str, date], Decimal] = {}
    for row in rows:
        key = (row.person_id, row.business_date)
        result[key] = result.get(key, Decimal("0")) + Decimal(str(row.amount))
    return result


def render_store_export_at_revision(
    session: Session,
    *,
    tenant_id: str,
    month: Month,
    store_id: str,
    revision: int,
) -> ExportEnvelope:
    """Render one store from exactly one calendar-derived revision."""

    store = session.execute(
        select(Store).where(Store.tenant_id == tenant_id, Store.id == store_id)
    ).scalar_one_or_none()
    if store is None:
        raise ValueError("STORE_NOT_FOUND")

    assignments = _assignments(
        session,
        tenant_id=tenant_id,
        month_id=month.id,
        revision=revision,
        store_ids=[store_id],
    )
    working_person_ids = {
        row.person_id for row in assignments if row.status == "WORKING"
    }
    persons = _people(
        session,
        tenant_id=tenant_id,
        person_ids=working_person_ids,
    )
    agents = persons[:2]
    persons_by_id = {person.id: person for person in persons}
    assignments_by_person_day = {
        (row.person_id, row.business_date): row for row in assignments
    }
    pontaj_rows = _pontaj(
        session,
        tenant_id=tenant_id,
        month_id=month.id,
        revision=revision,
        person_ids=working_person_ids,
    )
    grid_rows = list(
        session.execute(
            select(GridCalculation).where(
                GridCalculation.tenant_id == tenant_id,
                GridCalculation.month_id == month.id,
                GridCalculation.store_id == store_id,
                GridCalculation.revision == revision,
            )
        ).scalars()
    )
    grid_by_person = {row.person_id: row for row in grid_rows}
    sales_by_person_day = _sales(
        session,
        tenant_id=tenant_id,
        month_id=month.id,
        store_id=store_id,
        revision=revision,
    )

    # The established formatting helper displays ``month.revision`` in the
    # workbook header. Supply a read-only view whose revision is the pinned data
    # revision rather than a later administrative close/reopen revision.
    month_view = cast(
        Month,
        SimpleNamespace(
            id=month.id,
            year=month.year,
            month=month.month,
            revision=revision,
        ),
    )

    workbook = Workbook()
    grila = workbook.active
    if grila is None:
        raise RuntimeError("Workbook has no active sheet")
    _write_grila_tab(
        grila,
        month=month_view,
        store=store,
        agents=agents,
        grid_by_person=grid_by_person,
        assignments_by_person_day=assignments_by_person_day,
        sales_by_person_day=sales_by_person_day,
        holiday_labels={},
        month_year=month.year,
        month_month=month.month,
    )
    pontaj_sheet = workbook.create_sheet("Pontaj")
    _write_pontaj_tab(
        pontaj_sheet,
        pontaj_rows=pontaj_rows,
        assignments_by_person_day=assignments_by_person_day,
        persons_by_id=persons_by_id,
        month_year=month.year,
        month_month=month.month,
    )
    payload = save_workbook_deterministic(workbook)
    filename = _safe_filename(store.internal_code, month_view, "grila_pontaj")
    checksum = _checksum(payload)
    return ExportEnvelope(
        bytes_=payload,
        filename=filename,
        checksum=checksum,
        summary={
            "schema": SCHEMA,
            "tenant_id": tenant_id,
            "month_id": month.id,
            "store_id": store_id,
            "revision": revision,
            "filename": filename,
            "checksum_sha256": checksum,
            "kind": "EXPORT_XLSX_STORE",
            "rows_grid": len(grid_rows),
            "rows_pontaj": len(pontaj_rows),
            "agents_count": len(agents),
        },
    )


def _selected_store_ids(
    session: Session,
    *,
    tenant_id: str,
    store_ids: Iterable[str] | None,
) -> list[str]:
    query = select(Store).where(
        Store.tenant_id == tenant_id,
        Store.is_active.is_(True),
    )
    if store_ids is not None:
        query = query.where(Store.id.in_(list(store_ids)))
    stores = list(session.execute(query.order_by(Store.internal_code)).scalars())
    return [store.id for store in stores]


def render_pontaj_only_export_at_revision(
    session: Session,
    *,
    tenant_id: str,
    month: Month,
    revision: int,
    store_ids: Iterable[str] | None = None,
) -> ExportEnvelope:
    """Render Pontaj from one revision across the requested store set."""

    selected = _selected_store_ids(
        session,
        tenant_id=tenant_id,
        store_ids=store_ids,
    )
    assignments = _assignments(
        session,
        tenant_id=tenant_id,
        month_id=month.id,
        revision=revision,
        store_ids=selected,
    )
    person_ids = {row.person_id for row in assignments if row.status == "WORKING"}
    pontaj_rows = _pontaj(
        session,
        tenant_id=tenant_id,
        month_id=month.id,
        revision=revision,
        person_ids=person_ids,
    )
    persons = _people(session, tenant_id=tenant_id, person_ids=person_ids)
    assignments_by_person_day = {
        (row.person_id, row.business_date): row for row in assignments
    }

    workbook = Workbook()
    sheet = workbook.active
    if sheet is None:
        raise RuntimeError("Workbook has no active sheet")
    _write_pontaj_tab(
        sheet,
        pontaj_rows=pontaj_rows,
        assignments_by_person_day=assignments_by_person_day,
        persons_by_id={person.id: person for person in persons},
        month_year=month.year,
        month_month=month.month,
    )
    payload = save_workbook_deterministic(workbook)
    suffix = "pontaj_all" if len(selected) != 1 else "pontaj_1"
    filename = f"ugrile_{month.year:04d}-{month.month:02d}_{suffix}.xlsx"
    checksum = _checksum(payload)
    return ExportEnvelope(
        bytes_=payload,
        filename=filename,
        checksum=checksum,
        summary={
            "schema": SCHEMA,
            "tenant_id": tenant_id,
            "month_id": month.id,
            "revision": revision,
            "filename": filename,
            "checksum_sha256": checksum,
            "kind": "EXPORT_PONTAJ_ONLY",
            "rows_pontaj": len(pontaj_rows),
            "stores_included": selected,
        },
    )


def render_bulk_export_at_revision(
    session: Session,
    *,
    tenant_id: str,
    month: Month,
    revision: int,
    store_ids: Iterable[str] | None = None,
) -> ExportEnvelope:
    """Render a ZIP whose every workbook is pinned to the same revision."""

    selected = _selected_store_ids(
        session,
        tenant_id=tenant_id,
        store_ids=store_ids,
    )
    if not selected:
        raise ValueError("NO_STORES")
    stores = {
        store.id: store
        for store in session.execute(
            select(Store).where(
                Store.tenant_id == tenant_id,
                Store.id.in_(selected),
            )
        ).scalars()
    }
    generation = _determine_generation(
        session,
        tenant_id=tenant_id,
        month_year=month.year,
        month_month=month.month,
    )

    entries: list[dict[str, object]] = []
    archive_entries: list[tuple[str, bytes]] = []
    for store_id in selected:
        envelope = render_store_export_at_revision(
            session,
            tenant_id=tenant_id,
            month=month,
            store_id=store_id,
            revision=revision,
        )
        store = stores[store_id]
        archive_entries.append((envelope.filename, envelope.bytes_))
        entries.append(
            {
                "store_id": store.id,
                "internal_code": store.internal_code,
                "filename": envelope.filename,
                "checksum_sha256": envelope.checksum,
                "size_bytes": len(envelope.bytes_),
                "kind": "EXPORT_XLSX_STORE",
            }
        )

    from ..domain.rule_pack import RULE_PACK_VERSION

    manifest = {
        "schema": SCHEMA,
        "tenant_id": tenant_id,
        "month_id": month.id,
        "year": month.year,
        "month": month.month,
        "revision": revision,
        "rule_pack_version": RULE_PACK_VERSION,
        "generation": generation,
        "store_count": len(entries),
        "entries": entries,
    }
    archive_entries.append(
        (
            "manifest.json",
            json.dumps(
                manifest,
                sort_keys=True,
                ensure_ascii=False,
                indent=2,
            ).encode("utf-8"),
        )
    )
    payload = deterministic_zip(archive_entries)
    filename = f"ugrile_{month.year:04d}-{month.month:02d}_bulk.zip"
    checksum = _checksum(payload)
    return ExportEnvelope(
        bytes_=payload,
        filename=filename,
        checksum=checksum,
        summary={
            "schema": SCHEMA,
            "tenant_id": tenant_id,
            "month_id": month.id,
            "revision": revision,
            "filename": filename,
            "checksum_sha256": checksum,
            "kind": "EXPORT_XLSX_BULK",
            "store_count": len(entries),
            "rule_pack_version": manifest["rule_pack_version"],
            "generation": manifest["generation"],
        },
    )


__all__ = [
    "render_bulk_export_at_revision",
    "render_pontaj_only_export_at_revision",
    "render_store_export_at_revision",
]
