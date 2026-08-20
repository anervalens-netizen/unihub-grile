"""Safe, deterministic XLSX schedule template and preview parser.

The workbook is bound to a server-issued, single-use import contract (see
:mod:`ugrile.services.schedule_contract`). The opaque token travels inside
the hidden, sheet-protected ``Manifest`` tab and is never used as the
authority itself: the server re-reads the catalog, the effective manager
scope and the month revision at validation time and compares them against
the persisted contract row.

Cell encoding:

* ``BLOCAT`` is a locked cell produced for a person whose home store is
  outside the manager's effective scope on that date. It parses back as
  *no change* (``None``) so untouched blocked cells never create writes.
* ``LIBER`` / ``CONCEDIU`` map to ``OFF`` / ``LEAVE``.
* ``NORMAL - <code>``, ``SUPLIMENTAR ACASĂ - <code>`` and
  ``SUPLIMENTAR - <code>`` map to working kinds on the given store.

Formula cells and merged day cells are rejected as tampering rather than
silently decoded (an empty merged cell would otherwise read as ``OFF``).
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from ..domain.enums import DayStatus, WorkingKind
from .calendar import CalendarChange

SCHEMA = "UGRILE-S2-SCHEDULE-V2"

DAY_FIRST_COLUMN = 4  # column D = day 1
DAY_LAST_COLUMN = 34  # column AH = day 31


@dataclass(frozen=True, slots=True)
class ParsedSchedule:
    tenant_id: str
    month_id: str
    base_revision: int
    contract_token: str
    changes: list[CalendarChange]
    errors: list[dict[str, Any]]
    warnings: list[dict[str, Any]]


def _decode(
    value: object, person_id: str, day: date, stores: dict[str, str]
) -> CalendarChange | None:
    if value is None or str(value).strip() == "":
        return CalendarChange(person_id, day, None, DayStatus.OFF)
    text = str(value).strip()
    if text == "BLOCAT":
        return None
    if text == "LIBER":
        return CalendarChange(person_id, day, None, DayStatus.OFF)
    if text == "CONCEDIU":
        return CalendarChange(person_id, day, None, DayStatus.LEAVE)
    for prefix, status in (
        ("NORMAL - ", WorkingKind.NORMAL),
        ("SUPLIMENTAR ACASĂ - ", WorkingKind.EXTRA_HOME),
        ("SUPLIMENTAR - ", WorkingKind.EXTRA_OTHER),
    ):
        if text.startswith(prefix):
            code = text[len(prefix) :].strip()
            if code not in stores:
                raise ValueError("UNKNOWN_STORE")
            return CalendarChange(person_id, day, stores[code], DayStatus.WORKING, status)
    raise ValueError("MALFORMED_CELL")


def _encode(change: CalendarChange | None, stores_by_id: dict[str, str]) -> str:
    if change is None or change.status == DayStatus.OFF:
        return "LIBER"
    if change.status == DayStatus.LEAVE:
        return "CONCEDIU"
    if change.store_id is None or change.working_kind is None:
        return "LIBER"
    code = stores_by_id.get(change.store_id)
    if code is None:
        return "LIBER"
    prefix = {
        WorkingKind.NORMAL: "NORMAL",
        WorkingKind.EXTRA_HOME: "SUPLIMENTAR ACASĂ",
        WorkingKind.EXTRA_OTHER: "SUPLIMENTAR",
    }[change.working_kind]
    return f"{prefix} - {code}"


def _day_values(
    person: dict[str, str],
    year: int,
    month: int,
    calendar_by_key: dict[tuple[str, date], CalendarChange],
    stores_by_id: dict[str, str],
    store_ids_by_code: dict[str, str],
    allowed_store_ids_by_date: dict[date, set[str]] | None,
) -> list[str]:
    """Render the 31 day cells for one person row.

    When ``allowed_store_ids_by_date`` is provided (date-specific scope), any
    day on which the person's home store is outside the scope renders as the
    locked ``BLOCAT`` marker instead of an editable value. In-scope days keep
    their current calendar value so a partial-month round trip never loses the
    existing schedule.
    """

    values: list[str] = []
    person_id = person["person_id"]
    home_store_id = store_ids_by_code.get(person.get("home_store_code", ""))
    for day in range(1, 32):
        try:
            business_date = date(year, month, day)
        except ValueError:
            values.append("BLOCAT" if allowed_store_ids_by_date is not None else "LIBER")
            continue
        if (
            allowed_store_ids_by_date is not None
            and home_store_id not in allowed_store_ids_by_date.get(business_date, set())
        ):
            values.append("BLOCAT")
        else:
            values.append(_encode(calendar_by_key.get((person_id, business_date)), stores_by_id))
    return values


def _allowed_for_person(
    person: dict[str, str],
    stores: dict[str, str],
    allowed_store_ids: set[str] | None = None,
) -> list[str]:
    """Dropdown options for one person, constrained to the writable stores.

    ``allowed_store_ids`` is the union of the manager's effective scope across
    the month; individual cells are still re-validated per date at apply time.
    """

    visible_stores = {
        code: store_id
        for code, store_id in stores.items()
        if allowed_store_ids is None or store_id in allowed_store_ids
    }
    home_store = person.get("home_store_code", "")
    values = ["LIBER", "CONCEDIU"]
    if home_store in visible_stores:
        values.extend([f"NORMAL - {home_store}", f"SUPLIMENTAR ACASĂ - {home_store}"])
    values.extend(f"SUPLIMENTAR - {code}" for code in visible_stores if code != home_store)
    return values


def build_template(
    *,
    tenant_id: str,
    month_id: str,
    year: int,
    month: int,
    base_revision: int,
    contract_token: str,
    people: list[dict[str, str]],
    stores: dict[str, str],
    calendar: list[CalendarChange] | None = None,
    allowed_store_ids_by_date: dict[date, set[str]] | None = None,
) -> bytes:
    """Render the protected, contract-bound schedule workbook.

    The opaque ``contract_token`` is embedded in the hidden ``Manifest`` tab.
    The manifest, the ``_Lists`` tab and every manager sheet are sheet
    protected; only the day cells are unlocked. The manifest protection does
    not make the workbook tamper-proof on its own — the server-side contract
    validation is the authority.
    """

    calendar_by_key = {
        (change.person_id, change.business_date): change for change in (calendar or [])
    }
    store_codes_by_id = {store_id: code for code, store_id in stores.items()}
    allowed_store_ids = (
        set().union(*allowed_store_ids_by_date.values()) if allowed_store_ids_by_date else None
    )
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("workbook has no active sheet")
    ws.title = "Instrucțiuni"
    ws.append(["schema", SCHEMA])
    ws.append(["tenant_id", tenant_id])
    ws.append(["month_id", month_id])
    ws.append(["base_revision", base_revision])
    ws.append(
        ["legend", "LIBER | CONCEDIU | NORMAL - cod | SUPLIMENTAR ACASĂ - cod | SUPLIMENTAR - cod"]
    )
    ws.append(
        [
            "info",
            "BLOCAT = outside your effective scope on that date; leave it untouched.",
        ]
    )
    lists = wb.create_sheet("_Lists")
    lists.sheet_state = "hidden"
    lists.protection.sheet = True
    dropdown_ranges: dict[tuple[str, ...], str] = {}
    next_list_column = 1

    def dropdown_range(values: list[str]) -> str:
        nonlocal next_list_column
        key = tuple(values)
        existing = dropdown_ranges.get(key)
        if existing is not None:
            return existing
        column = next_list_column
        next_list_column += 1
        for row, value in enumerate(values, start=1):
            lists.cell(row, column, value)
        letter = get_column_letter(column)
        formula = f"='_Lists'!${letter}$1:${letter}${len(values)}"
        dropdown_ranges[key] = formula
        return formula

    used_sheet_names: set[str] = set(wb.sheetnames)
    for manager, rows in _by_manager(people):
        sheet_name = _safe_sheet_name(manager, used_sheet_names)
        used_sheet_names.add(sheet_name)
        tab = wb.create_sheet(sheet_name)
        headers = ["person_id", "nume", "magazin_bază"] + [str(d) for d in range(1, 32)]
        tab.append(headers)
        for person in rows:
            day_values = _day_values(
                person,
                year,
                month,
                calendar_by_key,
                store_codes_by_id,
                stores,
                allowed_store_ids_by_date,
            )
            tab.append(
                [
                    person["person_id"],
                    person.get("display_name", ""),
                    person.get("home_store_code", ""),
                    *day_values,
                ]
            )
            row_number = tab.max_row
            if allowed_store_ids_by_date is None:
                dv = DataValidation(
                    type="list",
                    formula1=dropdown_range(_allowed_for_person(person, stores, allowed_store_ids)),
                    allow_blank=False,
                )
                tab.add_data_validation(dv)
                dv.add(f"D{row_number}:AH{row_number}")
            else:
                for day_number, value in enumerate(day_values, start=1):
                    try:
                        business_date = date(year, month, day_number)
                    except ValueError:
                        choices = ["BLOCAT"]
                    else:
                        if value == "BLOCAT":
                            choices = ["BLOCAT"]
                        else:
                            choices = _allowed_for_person(
                                person, stores, allowed_store_ids_by_date[business_date]
                            )
                    dv = DataValidation(
                        type="list",
                        formula1=dropdown_range(choices),
                        allow_blank=False,
                    )
                    tab.add_data_validation(dv)
                    dv.add(tab.cell(row_number, day_number + DAY_FIRST_COLUMN - 1).coordinate)
            for offset, value in enumerate(day_values, start=DAY_FIRST_COLUMN):
                if value != "BLOCAT":
                    tab.cell(row_number, offset).protection = Protection(locked=False)
        tab.column_dimensions["A"].hidden = True
        tab.freeze_panes = "D2"
        tab.protection.sheet = True
    manifest = wb.create_sheet("Manifest")
    manifest.sheet_state = "hidden"
    manifest.protection.sheet = True
    manifest.append(["schema", SCHEMA])
    manifest.append(["tenant_id", tenant_id])
    manifest.append(["month_id", month_id])
    manifest.append(["base_revision", base_revision])
    manifest.append(["contract_token", contract_token])
    manifest.append(["store_codes", "|".join(sorted(stores))])
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _safe_sheet_name(manager: str, used: set[str]) -> str:
    base = re.sub(r"[\[\]:*?/\\]", "_", manager).strip()[:31] or "Manager"
    candidate = base
    suffix = 2
    while candidate in used:
        tail = f"_{suffix}"
        candidate = f"{base[: 31 - len(tail)]}{tail}"
        suffix += 1
    return candidate


def _by_manager(people: list[dict[str, str]]) -> list[tuple[str, list[dict[str, str]]]]:
    grouped: dict[str, list[dict[str, str]]] = {}
    for p in people:
        grouped.setdefault(p.get("manager_code", "manager"), []).append(p)
    return sorted(grouped.items())


def _is_merged(worksheet: Any, coordinate: str) -> bool:
    return any(coordinate in rng for rng in worksheet.merged_cells.ranges)


def parse_schedule(
    data: bytes,
    *,
    expected_tenant_id: str,
    expected_month_id: str,
    year: int,
    month: int,
    stores: dict[str, str],
    known_person_ids: set[str] | None = None,
) -> ParsedSchedule:
    errors: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []
    changes: list[CalendarChange] = []
    seen_person_days: set[tuple[str, date]] = set()
    try:
        # data_only=False so formula cells are visible as such and rejected
        # instead of being silently decoded from a cached value (or None).
        wb = load_workbook(BytesIO(data), data_only=False, keep_links=False)
    except Exception as exc:
        return ParsedSchedule(
            tenant_id=expected_tenant_id,
            month_id=expected_month_id,
            base_revision=-1,
            contract_token="",
            changes=[],
            errors=[{"code": "INVALID_XLSX", "message": str(exc)}],
            warnings=[],
        )
    if "Instrucțiuni" not in wb.sheetnames or "Manifest" not in wb.sheetnames:
        return ParsedSchedule(
            tenant_id=expected_tenant_id,
            month_id=expected_month_id,
            base_revision=-1,
            contract_token="",
            changes=[],
            errors=[{"code": "INVALID_STRUCTURE"}],
            warnings=[],
        )
    manifest = wb["Manifest"]
    values = {
        str(manifest.cell(r, 1).value): manifest.cell(r, 2).value
        for r in range(1, manifest.max_row + 1)
    }
    tenant_id = str(values.get("tenant_id", ""))
    month_id = str(values.get("month_id", ""))
    contract_token = str(values.get("contract_token", "") or "")
    try:
        revision = int(str(values.get("base_revision", -1)))
    except (TypeError, ValueError):
        revision = -1
    if values.get("schema") != SCHEMA:
        errors.append({"code": "SCHEMA_MISMATCH"})
    if tenant_id != expected_tenant_id:
        errors.append({"code": "TENANT_MISMATCH"})
    if month_id != expected_month_id:
        errors.append({"code": "MONTH_MISMATCH"})
    try:
        for sheet in wb.worksheets:
            if sheet.title in {"Instrucțiuni", "Manifest", "_Lists"}:
                continue
            for row in range(2, sheet.max_row + 1):
                person_id = sheet.cell(row, 1).value
                if not person_id:
                    errors.append({"code": "MISSING_PERSON_ID", "sheet": sheet.title, "row": row})
                    continue
                if str(person_id).strip() != str(person_id):
                    errors.append({"code": "MALFORMED_PERSON_ID", "sheet": sheet.title, "row": row})
                if known_person_ids is not None and str(person_id) not in known_person_ids:
                    errors.append({"code": "UNKNOWN_PERSON", "sheet": sheet.title, "row": row})
                    continue
                for day_number in range(1, 32):
                    try:
                        day = date(year, month, day_number)
                    except ValueError:
                        continue
                    cell = sheet.cell(row, day_number + 3)
                    if cell.data_type == "f":
                        errors.append(
                            {
                                "code": "FORMULA_CELL",
                                "sheet": sheet.title,
                                "row": row,
                                "day": day_number,
                            }
                        )
                        continue
                    if _is_merged(sheet, cell.coordinate):
                        errors.append(
                            {
                                "code": "MERGED_CELL",
                                "sheet": sheet.title,
                                "row": row,
                                "day": day_number,
                            }
                        )
                        continue
                    try:
                        change = _decode(cell.value, str(person_id), day, stores)
                    except ValueError as exc:
                        errors.append(
                            {"code": str(exc), "sheet": sheet.title, "row": row, "day": day_number}
                        )
                        continue
                    if change:
                        key = (change.person_id, change.business_date)
                        if key in seen_person_days:
                            errors.append(
                                {
                                    "code": "DUPLICATE_PERSON_DAY",
                                    "sheet": sheet.title,
                                    "row": row,
                                    "day": day_number,
                                }
                            )
                        else:
                            seen_person_days.add(key)
                            changes.append(change)
    except Exception as exc:
        errors.append({"code": "PARSE_ERROR", "message": str(exc)})
    return ParsedSchedule(
        tenant_id=tenant_id,
        month_id=month_id,
        base_revision=revision,
        contract_token=contract_token,
        changes=changes,
        errors=errors,
        warnings=warnings,
    )


__all__ = [
    "ParsedSchedule",
    "SCHEMA",
    "build_template",
    "parse_schedule",
]
