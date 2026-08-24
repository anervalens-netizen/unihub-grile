"""XLSX export service (AC-14).

Renders three deterministic workbook kinds from the persisted read
models (PontajProjection, GridCalculation, SiteDayAssignment,
PontajHoursSnapshot semantics) following the contract documented in
``docs/MOBIUP_RULE_PACK.md §7`` (Pontaj standard) and the V2 Grila
card layout (2 agents per standard store, salary/projection cards,
calendar, holidays, supplementary).

Layout summary
--------------

``Grila`` tab — V2 card layout:

* Row 1-4: magazin header (nume, cod intern, luna, schema, revision).
* Row 6: card-1 header (Persoana 1 + components).
* Row 7: card-1 row (componente salariale + proiecție).
* Row 9: card-2 header (Persoana 2 + components).
* Row 10: card-2 row.
* Row 12: header calendar lunar (days 1..31 in coloanele C..AG).
* Row 13: card-1 calendar (assignment per day + working_kind badge).
* Row 14: card-2 calendar.

``Pontaj`` tab — standard Mobiup C8:AG31:

* Row 1: header (Persoana + day 1..31 in C..AG + Total ore in AH).
* Row 2: blank / spacer.
* Row 8, 11, 14, 17, 20, 23, 26, 29: per-block day rows (Net hours),
  followed by interval (row r+1) and pause (row r+2).
* Historical participants use the available block-start rows in deterministic
  person order, up to all eight blocks. Rendering fails closed above layout
  capacity; Pontaj never silently truncates a historical participant.
* Total at column AH per active block: ``AHr = SUM(Cr:AGr)``.

Per-store export filters assignments/pontaj to the store's own grid
cells; bulk and pontaj-only honour the same per-store scope and the
contract filter (firmă/manager/magazin from the request payload).

No external links, no live Google writes, deterministic SHA-256 checksums
in the bulk manifest.
"""

from __future__ import annotations

import hashlib
import io
import json
import zipfile
from collections import defaultdict
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import date, time, timedelta
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from ..domain.errors import ValidationError
from ..repositories.models import (
    ExportRun,
    GridCalculation,
    Month,
    Person,
    PontajProjection,
    SiteDayAssignment,
    Store,
)

SCHEMA = "UGRILE-S5-XLSX-V2"
MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Standard Pontaj block rows (8, 11, 14, 17, 20, 23, 26, 29).
PONTAJ_BLOCK_STARTS: tuple[int, ...] = (8, 11, 14, 17, 20, 23, 26, 29)

# Standard shift per docs/MOBIUP_RULE_PACK.md §7 for NORMAL/EXTRA_HOME/EXTRA_OTHER.
STANDARD_INTERVAL_START = time(10, 0)
STANDARD_INTERVAL_END = time(22, 0)
STANDARD_PAUSE_MINUTES = 60
STANDARD_NET_HOURS = Decimal("11")

# Column layout (1-indexed). Day 1 starts at column C = 3; the standard
# Pontaj contract places day 31 at column AG = 33, and column AH = 34 holds
# the total monthly hours.
DAY_1_COL = 3   # column C
DAY_31_COL = 33  # column AG (day 31)
TOTAL_COL = 34   # column AH (monthly total)
WEEKEND_FILL = PatternFill("solid", fgColor="FFE699")


@dataclass(frozen=True, slots=True)
class ExportEnvelope:
    bytes_: bytes
    filename: str
    checksum: str
    summary: dict[str, object]


# --- formatting helpers ---


def _romanian_money(value: Decimal | int | float | None) -> str:
    if value is None:
        return "0,00"
    decimal = Decimal(str(value)).quantize(Decimal("0.01"))
    s = format(decimal, "f")
    int_part, _, frac_part = s.partition(".")
    sign = ""
    if int_part.startswith("-"):
        sign = "-"
        int_part = int_part[1:]
    grouped = ""
    while len(int_part) > 3:
        grouped = "." + int_part[-3:] + grouped
        int_part = int_part[:-3]
    grouped = int_part + grouped
    return f"{sign}{grouped},{frac_part}"


def _hours(value: Decimal | None) -> str:
    if value is None:
        return "0"
    return format(Decimal(str(value)).quantize(Decimal("0.01")), "f")


def _safe_filename(store_code: str, month: Month, suffix: str) -> str:
    return (
        f"ugrile_{month.year:04d}-{month.month:02d}_"
        f"{store_code.replace('/', '_').replace(' ', '_')}_{suffix}.xlsx"
    )


def _checksum(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _header_fill() -> PatternFill:
    return PatternFill("solid", fgColor="1F5FBF")


def _header_font() -> Font:
    return Font(bold=True, color="FFFFFF")


def _thin_border() -> Border:
    side = Side(style="thin", color="D4D7DD")
    return Border(left=side, right=side, top=side, bottom=side)


# --- Grila tab (V2 cards layout) ---


def _grila_agents_for_store(
    session: Session,
    *,
    tenant_id: str,
    store_id: str,
    month_id: str,
) -> list[Person]:
    """Return the two agents the V2 Grila card layout expects.

    The standard two-agent store uses the first two distinct persons
    that have at least one WORKING assignment on this store/month,
    sorted by internal_code for determinism. Rows 13 and 14 (calendar)
    and rows 6-7 + 9-10 (cards) are reserved only for these two agents.
    """
    rows = list(
        session.execute(
            select(SiteDayAssignment.person_id)
            .where(
                SiteDayAssignment.tenant_id == tenant_id,
                SiteDayAssignment.month_id == month_id,
                SiteDayAssignment.store_id == store_id,
                SiteDayAssignment.status == "WORKING",
            )
            .distinct()
        ).scalars()
    )
    if not rows:
        return []
    persons = list(
        session.execute(
            select(Person).where(
                Person.tenant_id == tenant_id, Person.id.in_(rows)
            )
        ).scalars()
    )
    persons.sort(key=lambda p: (p.internal_code or "", p.id))
    return persons[:2]


def _write_grila_tab(
    ws: Worksheet,
    *,
    month: Month,
    store: Store,
    agents: list[Person],
    grid_by_person: dict[str, GridCalculation],
    assignments_by_person_day: dict[tuple[str, date], SiteDayAssignment],
    sales_by_person_day: dict[tuple[str, date], Decimal],
    holiday_labels: dict[date, str],
    month_year: int,
    month_month: int,
) -> None:
    ws.title = "Grila"
    days_in_month = (date(month_year, month_month % 12 + 1, 1) - timedelta(days=1)).day if month_month < 12 else 31
    days_in_month = (date(month_year + (1 if month_month == 12 else 0), 1 if month_month == 12 else month_month + 1, 1) - timedelta(days=1)).day
    # Header block
    ws["A1"] = f"Magazin: {store.name}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Cod intern: {store.internal_code}"
    ws["A3"] = f"Companie: {store.company_code}"
    ws["A4"] = (
        f"Luna: {month_year:04d}-{month_month:02d}  "
        f"Revizie: {month.revision}  Schema: {SCHEMA}"
    )
    # Two-agent cards (rows 6..10), one card per agent.
    card_header_fill = _header_fill()
    card_header_font = _header_font()
    card_columns = [
        "Persoana",
        "Acord global",
        "Salariu fix",
        "Tichete",
        "Comision principal",
        "Bonus",
        "Plata fixa extra",
        "Comision EXTRA_OTHER",
        "Comision SIM",
        "Comision E-pay",
        "Incentive",
        "Flip",
        "Total salariu",
        "Salariu cash",
    ]
    for card_idx in range(2):
        if card_idx >= len(agents):
            break
        header_row = 6 + card_idx * 3
        data_row = header_row + 1
        ws.cell(row=header_row, column=1, value=f"Card #{card_idx + 1}").font = Font(bold=True)
        for col, header in enumerate(card_columns, start=1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.fill = card_header_fill
            cell.font = card_header_font
            cell.border = _thin_border()
        agent = agents[card_idx]
        grid = grid_by_person.get(agent.id)
        ws.cell(row=data_row, column=1, value=agent.display_name or agent.internal_code).border = _thin_border()
        if grid is None:
            for col in range(2, 1 + len(card_columns)):
                ws.cell(row=data_row, column=col, value="").border = _thin_border()
        else:
            payload = json.loads(grid.payload or "{}")
            components = payload.get("components", {}) if isinstance(payload, dict) else {}
            values = [
                components.get("acord_global", "Acord curent"),
                _romanian_money(components.get("salary")),
                _romanian_money(components.get("tickets")),
                _romanian_money(components.get("main_commission")),
                _romanian_money(components.get("main_bonus")),
                _romanian_money(components.get("extra_fixed_pay")),
                _romanian_money(components.get("extra_other_commission")),
                _romanian_money(components.get("sim_commission")),
                _romanian_money(components.get("epay_commission")),
                _romanian_money(components.get("incentive")),
                _romanian_money(components.get("flip")),
                _romanian_money(components.get("total_salary")),
                _romanian_money(components.get("salary_cash")),
            ]
            for col, value in enumerate(values, start=2):
                cell = ws.cell(row=data_row, column=col, value=value)
                cell.border = _thin_border()
                cell.alignment = Alignment(horizontal="right")
    # Calendar block — rows 13..14, columns D..AG (day 1..31) with badges.
    cal_header_row = 12
    for col in range(1, 2):
        ws.cell(row=cal_header_row, column=col, value="Calendar").font = Font(bold=True)
    for day in range(1, 32):
        cell = ws.cell(row=cal_header_row, column=DAY_1_COL + day - 1, value=day)
        cell.fill = _header_fill()
        cell.font = _header_font()
        cell.border = _thin_border()
    for card_idx in range(2):
        if card_idx >= len(agents):
            break
        row = 13 + card_idx
        agent = agents[card_idx]
        ws.cell(row=row, column=1, value=agent.display_name or agent.internal_code).border = _thin_border()
        for day in range(1, 32):
            target = date(month_year, month_month, day) if day <= days_in_month else None
            day_cell = ws.cell(row=row, column=DAY_1_COL + day - 1)
            day_cell.border = _thin_border()
            day_cell.alignment = Alignment(horizontal="center")
            if target is None:
                continue
            assignment = assignments_by_person_day.get((agent.id, target))
            if assignment is None or assignment.status != "WORKING":
                day_cell.value = ""
                continue
            sales = sales_by_person_day.get((agent.id, target), Decimal("0"))
            badge = assignment.working_kind or "NORMAL"
            day_cell.value = f"{badge[:3]}\n{int(sales)}" if sales else badge[:3]
    # Holiday markers (informational; non-blocking per docs/MOBIUP_RULE_PACK.md §9).
    if holiday_labels:
        start_row = 16
        ws.cell(row=start_row, column=1, value="Sarbatori legale (marker informativ):").font = Font(italic=True)
        for offset, (d, label) in enumerate(sorted(holiday_labels.items()), start=1):
            ws.cell(row=start_row + offset, column=1, value=f"{d.strftime('%d/%m/%Y')}: {label}").font = Font(italic=True)


# --- Pontaj tab (standard C8:AG31, AH total) ---


def _write_pontaj_tab(
    ws: Worksheet,
    *,
    pontaj_rows: list[PontajProjection],
    assignments_by_person_day: dict[tuple[str, date], SiteDayAssignment],
    persons_by_id: dict[str, Person],
    month_year: int,
    month_month: int,
    active_block_rows: Iterable[int] = PONTAJ_BLOCK_STARTS,
) -> None:
    """Write the standard Mobiup Pontaj tab.

    Layout:
      row 1: Persoana + day 1..31 + Total ore (AH)
      row 2..7: header band / metadata
      rows 8, 11, 14, 17, 20, 23, 26, 29: per-agent block starts;
        row r = net hours, r+1 = interval, r+2 = pause.
      column AH = total net hours per row r.
    """
    ws.title = "Pontaj"
    days_in_month = (date(month_year + (1 if month_month == 12 else 0), 1 if month_month == 12 else month_month + 1, 1) - timedelta(days=1)).day
    # Row 1 header.
    ws.cell(row=1, column=1, value="Persoana")
    for day in range(1, 32):
        cell = ws.cell(row=1, column=DAY_1_COL + day - 1, value=day)
        cell.fill = _header_fill()
        cell.font = _header_font()
        cell.border = _thin_border()
    total_col = ws.cell(row=1, column=TOTAL_COL, value="Total ore (AH)")
    total_col.fill = _header_fill()
    total_col.font = _header_font()
    total_col.border = _thin_border()
    # Bucket pontaj rows per person.
    by_person: dict[str, dict[date, PontajProjection]] = defaultdict(dict)
    for row in pontaj_rows:
        by_person[row.person_id][row.business_date] = row
    # The standard sheet provides eight 3-row participant blocks. Historical
    # people are part of payroll truth even after transfer/deactivation, so render
    # every person supplied by the revision-pinned selector. Never turn overflow
    # into a plausible but incomplete workbook.
    ordered_block_rows = list(active_block_rows)
    sorted_person_ids = sorted(
        persons_by_id.keys(),
        key=lambda pid: ((persons_by_id[pid].internal_code or ""), pid),
    )
    if len(sorted_person_ids) > len(ordered_block_rows):
        raise ValidationError(
            "Pontaj layout capacity exceeded",
            details={
                "code": "PONTAJ_LAYOUT_CAPACITY_EXCEEDED",
                "capacity": len(ordered_block_rows),
                "participants": len(sorted_person_ids),
            },
        )
    block_persons = sorted_person_ids
    for block_idx, start_row in enumerate(ordered_block_rows):
        if block_idx >= len(block_persons):
            break
        person_id = block_persons[block_idx]
        person = persons_by_id.get(person_id)
        ws.cell(
            row=start_row,
            column=1,
            value=person.display_name if person is not None else person_id,
        ).border = _thin_border()
        interval_cell = ws.cell(
            row=start_row + 1,
            column=1,
            value=f"Interval: {STANDARD_INTERVAL_START.strftime('%H:%M')}-{STANDARD_INTERVAL_END.strftime('%H:%M')}",
        )
        interval_cell.font = Font(italic=True, size=9)
        interval_cell.border = _thin_border()
        pause_cell = ws.cell(
            row=start_row + 2,
            column=1,
            value=f"Pauza: {STANDARD_PAUSE_MINUTES} min",
        )
        pause_cell.font = Font(italic=True, size=9)
        pause_cell.border = _thin_border()
        total = Decimal("0")
        for day in range(1, 32):
            target = date(month_year, month_month, day) if day <= days_in_month else None
            day_cell = ws.cell(row=start_row, column=DAY_1_COL + day - 1)
            day_cell.border = _thin_border()
            day_cell.alignment = Alignment(horizontal="right")
            if target is None:
                continue
            assignment = assignments_by_person_day.get((person_id, target))
            projection = by_person.get(person_id, {}).get(target)
            is_working = (
                assignment is not None
                and assignment.status == "WORKING"
                and projection is not None
                and projection.status == "WORKING"
            )
            if is_working and projection is not None:
                day_cell.value = _hours(projection.hours)
                total += Decimal(str(projection.hours))
            else:
                day_cell.value = ""
            # Weekend fill per docs/MOBIUP_RULE_PACK.md §7.
            if target is not None and target.weekday() >= 5:
                day_cell.fill = WEEKEND_FILL
        total_cell = ws.cell(row=start_row, column=TOTAL_COL, value=_hours(total))
        total_cell.border = _thin_border()
        total_cell.alignment = Alignment(horizontal="right")
        total_cell.font = Font(bold=True)


# --- selection helpers ---


def _pontaj_rows_for_store(
    session: Session,
    *,
    tenant_id: str,
    store_id: str,
    month_id: str,
) -> list[PontajProjection]:
    """Pontaj rows whose person has at least one WORKING assignment on
    ``store_id`` in ``month_id``. Strict per-store filter."""
    assigned_person_ids = list(
        session.execute(
            select(SiteDayAssignment.person_id)
            .distinct()
            .where(
                SiteDayAssignment.tenant_id == tenant_id,
                SiteDayAssignment.month_id == month_id,
                SiteDayAssignment.store_id == store_id,
                SiteDayAssignment.status == "WORKING",
            )
        ).scalars()
    )
    if not assigned_person_ids:
        return []
    return list(
        session.execute(
            select(PontajProjection).where(
                PontajProjection.tenant_id == tenant_id,
                PontajProjection.month_id == month_id,
                PontajProjection.person_id.in_(assigned_person_ids),
            )
        ).scalars()
    )


def _assignments_for_store(
    session: Session,
    *,
    tenant_id: str,
    store_id: str,
    month_id: str,
) -> list[SiteDayAssignment]:
    return list(
        session.execute(
            select(SiteDayAssignment).where(
                SiteDayAssignment.tenant_id == tenant_id,
                SiteDayAssignment.month_id == month_id,
                SiteDayAssignment.store_id == store_id,
            ).order_by(SiteDayAssignment.person_id, SiteDayAssignment.business_date)
        ).scalars()
    )


def _sales_for_store(
    session: Session,
    *,
    tenant_id: str,
    store_id: str,
    month_id: str,
    month_year: int,
    month_month: int,
) -> dict[tuple[str, date], Decimal]:
    """Sum attributed sales per (person, day) for the store, derived from
    SalesPersonDayProjection at the current month revision."""
    month = session.execute(select(Month).where(Month.id == month_id)).scalar_one()
    rows = list(
        session.execute(
            select(
                __import__("ugrile.repositories.models", fromlist=["SalesPersonDayProjection"]).SalesPersonDayProjection
            ).where(
                __import__("ugrile.repositories.models", fromlist=["SalesPersonDayProjection"]).SalesPersonDayProjection.tenant_id == tenant_id,
                __import__("ugrile.repositories.models", fromlist=["SalesPersonDayProjection"]).SalesPersonDayProjection.month_id == month_id,
                __import__("ugrile.repositories.models", fromlist=["SalesPersonDayProjection"]).SalesPersonDayProjection.store_id == store_id,
                __import__("ugrile.repositories.models", fromlist=["SalesPersonDayProjection"]).SalesPersonDayProjection.revision == month.revision,
            )
        ).scalars()
    )
    out: dict[tuple[str, date], Decimal] = {}
    for row in rows:
        key = (row.person_id, row.business_date)
        out[key] = out.get(key, Decimal("0")) + Decimal(str(row.amount))
    return out


def _determine_generation(session: Session, *, tenant_id: str, month_year: int, month_month: int) -> str:
    from ..repositories.models import SalesStoreDay

    rows = list(
        session.execute(
            select(func.distinct(SalesStoreDay.generation)).where(
                SalesStoreDay.tenant_id == tenant_id,
                func.extract("year", SalesStoreDay.business_date) == month_year,
                func.extract("month", SalesStoreDay.business_date) == month_month,
            )
        ).scalars()
    )
    if len(rows) == 1 and isinstance(rows[0], str):
        return rows[0]
    return "FIXTURE_V1"


# --- public renderers ---


def render_store_export(
    session: Session,
    *,
    tenant_id: str,
    month: Month,
    store_id: str,
) -> ExportEnvelope:
    """Render a per-magazin Grila + Pontaj workbook.

    Strict per-store filter: assignments, Pontaj rows, and GridCalculation
    rows are restricted to ``store_id``. The Pontaj tab uses the standard
    ``C8:AG31`` block layout with ``AH`` totals.
    """
    store = session.execute(
        select(Store).where(Store.tenant_id == tenant_id, Store.id == store_id)
    ).scalar_one_or_none()
    if store is None:
        raise ValueError("STORE_NOT_FOUND")

    grid_rows = list(
        session.execute(
            select(GridCalculation).where(
                GridCalculation.tenant_id == tenant_id,
                GridCalculation.month_id == month.id,
                GridCalculation.store_id == store_id,
            )
        ).scalars()
    )
    grid_by_person = {g.person_id: g for g in grid_rows}

    assignments = _assignments_for_store(
        session,
        tenant_id=tenant_id,
        store_id=store_id,
        month_id=month.id,
    )
    assignments_by_person_day = {
        (a.person_id, a.business_date): a for a in assignments
    }

    pontaj_rows = _pontaj_rows_for_store(
        session,
        tenant_id=tenant_id,
        store_id=store_id,
        month_id=month.id,
    )
    persons_in_store = list(
        session.execute(
            select(Person).where(
                Person.tenant_id == tenant_id,
                Person.id.in_({a.person_id for a in assignments if a.status == "WORKING"}),
            )
        ).scalars()
    )
    persons_by_id = {p.id: p for p in persons_in_store}
    agents = _grila_agents_for_store(
        session,
        tenant_id=tenant_id,
        store_id=store_id,
        month_id=month.id,
    )
    sales_by_person_day = _sales_for_store(
        session,
        tenant_id=tenant_id,
        store_id=store_id,
        month_id=month.id,
        month_year=month.year,
        month_month=month.month,
    )

    wb = Workbook()
    grila = wb.active
    if grila is None:
        raise RuntimeError("Workbook has no active sheet")
    _write_grila_tab(
        grila,
        month=month,
        store=store,
        agents=agents,
        grid_by_person=grid_by_person,
        assignments_by_person_day=assignments_by_person_day,
        sales_by_person_day=sales_by_person_day,
        holiday_labels={},
        month_year=month.year,
        month_month=month.month,
    )
    pontaj = wb.create_sheet("Pontaj")
    _write_pontaj_tab(
        pontaj,
        pontaj_rows=pontaj_rows,
        assignments_by_person_day=assignments_by_person_day,
        persons_by_id=persons_by_id,
        month_year=month.year,
        month_month=month.month,
    )
    buffer = io.BytesIO()
    wb.save(buffer)
    payload = buffer.getvalue()
    filename = _safe_filename(store.internal_code, month, "grila_pontaj")
    checksum = _checksum(payload)
    summary = {
        "schema": SCHEMA,
        "tenant_id": tenant_id,
        "month_id": month.id,
        "store_id": store_id,
        "filename": filename,
        "checksum_sha256": checksum,
        "kind": "EXPORT_XLSX_STORE",
        "rows_grid": len(grid_rows),
        "rows_pontaj": len(pontaj_rows),
        "agents_count": len(agents),
    }
    return ExportEnvelope(
        bytes_=payload,
        filename=filename,
        checksum=checksum,
        summary=summary,
    )


def render_pontaj_only_export(
    session: Session,
    *,
    tenant_id: str,
    month: Month,
    store_ids: Iterable[str] | None = None,
) -> ExportEnvelope:
    """Render a pontaj-only workbook spanning one or many stores.

    Each store contributes its assigned persons to the standard
    ``C8:AG31`` block layout. When ``store_ids`` is given, only those
    stores contribute; otherwise every active store in the tenant.
    """
    store_filter: list[str]
    if store_ids is not None:
        store_filter = [str(s) for s in store_ids]
    else:
        store_filter = [
            s.id
            for s in session.execute(
                select(Store).where(
                    Store.tenant_id == tenant_id, Store.is_active.is_(True)
                )
            ).scalars()
        ]
    all_assignments: list[SiteDayAssignment] = []
    all_pontaj: list[PontajProjection] = []
    for store_id in store_filter:
        all_assignments.extend(
            _assignments_for_store(
                session, tenant_id=tenant_id, store_id=store_id, month_id=month.id
            )
        )
        all_pontaj.extend(
            _pontaj_rows_for_store(
                session, tenant_id=tenant_id, store_id=store_id, month_id=month.id
            )
        )
    assignments_by_person_day = {
        (a.person_id, a.business_date): a for a in all_assignments
    }
    persons = list(
        session.execute(
            select(Person).where(
                Person.tenant_id == tenant_id,
                Person.id.in_({a.person_id for a in all_assignments if a.status == "WORKING"}),
            )
        ).scalars()
    )
    persons_by_id = {p.id: p for p in persons}

    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Workbook has no active sheet")
    _write_pontaj_tab(
        ws,
        pontaj_rows=all_pontaj,
        assignments_by_person_day=assignments_by_person_day,
        persons_by_id=persons_by_id,
        month_year=month.year,
        month_month=month.month,
    )
    buffer = io.BytesIO()
    wb.save(buffer)
    payload = buffer.getvalue()
    suffix = "pontaj_all" if not store_filter or len(store_filter) > 1 else f"pontaj_{len(store_filter)}"
    filename = f"ugrile_{month.year:04d}-{month.month:02d}_{suffix}.xlsx"
    checksum = _checksum(payload)
    summary = {
        "schema": SCHEMA,
        "tenant_id": tenant_id,
        "month_id": month.id,
        "filename": filename,
        "checksum_sha256": checksum,
        "kind": "EXPORT_PONTAJ_ONLY",
        "rows_pontaj": len(all_pontaj),
        "stores_included": list(store_filter),
    }
    return ExportEnvelope(
        bytes_=payload,
        filename=filename,
        checksum=checksum,
        summary=summary,
    )


def render_bulk_export(
    session: Session,
    *,
    tenant_id: str,
    month: Month,
    store_ids: Iterable[str] | None = None,
) -> ExportEnvelope:
    """Render a bulk ZIP: per-store XLSX + manifest.json with checksums."""
    store_query = select(Store).where(
        Store.tenant_id == tenant_id, Store.is_active.is_(True)
    )
    if store_ids is not None:
        store_query = store_query.where(Store.id.in_(list(store_ids)))
    stores = list(session.execute(store_query.order_by(Store.internal_code)).scalars())
    if not stores:
        raise ValueError("NO_STORES")
    generation = _determine_generation(
        session, tenant_id=tenant_id, month_year=month.year, month_month=month.month
    )

    entries: list[dict[str, object]] = []
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for store in stores:
            envelope = render_store_export(
                session, tenant_id=tenant_id, month=month, store_id=store.id
            )
            entry = {
                "store_id": store.id,
                "internal_code": store.internal_code,
                "filename": envelope.filename,
                "checksum_sha256": envelope.checksum,
                "size_bytes": len(envelope.bytes_),
                "kind": "EXPORT_XLSX_STORE",
            }
            zf.writestr(envelope.filename, envelope.bytes_)
            entries.append(entry)
        from ..domain.rule_pack import RULE_PACK_VERSION

        manifest = {
            "schema": SCHEMA,
            "tenant_id": tenant_id,
            "month_id": month.id,
            "year": month.year,
            "month": month.month,
            "revision": month.revision,
            "rule_pack_version": RULE_PACK_VERSION,
            "generation": generation,
            "store_count": len(entries),
            "entries": entries,
        }
        zf.writestr(
            "manifest.json",
            json.dumps(manifest, sort_keys=True, ensure_ascii=False, indent=2),
        )
    payload = buffer.getvalue()
    filename = f"ugrile_{month.year:04d}-{month.month:02d}_bulk.zip"
    checksum = _checksum(payload)
    summary = {
        "schema": SCHEMA,
        "tenant_id": tenant_id,
        "month_id": month.id,
        "filename": filename,
        "checksum_sha256": checksum,
        "kind": "EXPORT_XLSX_BULK",
        "store_count": len(entries),
        "rule_pack_version": manifest["rule_pack_version"],
        "generation": manifest["generation"],
    }
    return ExportEnvelope(
        bytes_=payload,
        filename=filename,
        checksum=checksum,
        summary=summary,
    )


def record_export_run(
    session: Session,
    *,
    tenant_id: str,
    kind: str,
    envelope: ExportEnvelope,
    artifact_uri: str,
) -> ExportRun:
    """Legacy entry point: insert a fresh DONE ExportRun.

    Prefer ``create_pending_export_run`` + ``finalize_export_run`` for the
    async API contract (the polling endpoint reads the ExportRun by id and
    the id must match the enqueue response).
    """

    row = ExportRun(
        tenant_id=tenant_id,
        kind=kind,
        status="DONE",
        summary=json.dumps(envelope.summary, sort_keys=True, ensure_ascii=False),
        artifact_uri=artifact_uri,
    )
    session.add(row)
    session.flush()
    return row


def create_pending_export_run(
    session: Session,
    *,
    tenant_id: str,
    kind: str,
    summary: dict[str, object] | None = None,
    artifact_uri_hint: str | None = None,
) -> ExportRun:
    """Pre-create a PENDING ``ExportRun`` for the async polling contract.

    The returned ``ExportRun.id`` is the stable identifier the API returns
    to the client on enqueue and that the polling endpoint
    (``GET /months/{id}/export/jobs/{job_id}``) reads back. The worker
    updates the same row to ``DONE``/``FAILED`` once the export completes
    via :func:`finalize_export_run`.
    """

    payload_summary: dict[str, object] = dict(summary or {})
    if artifact_uri_hint is not None:
        payload_summary.setdefault("artifact_uri_hint", artifact_uri_hint)
    row = ExportRun(
        tenant_id=tenant_id,
        kind=kind,
        status="PENDING",
        summary=json.dumps(payload_summary, sort_keys=True, ensure_ascii=False),
        artifact_uri=None,
    )
    session.add(row)
    session.flush()
    return row


def finalize_export_run(
    session: Session,
    *,
    export_run_id: int,
    tenant_id: str,
    status: str,
    artifact_uri: str | None = None,
    summary: dict[str, object] | None = None,
    errors: str | None = None,
) -> ExportRun | None:
    """Update an existing ``ExportRun`` to its terminal state.

    Returns the refreshed row, or ``None`` when the row is missing or owned
    by a different tenant (in which case the caller should treat the
    terminal status as authoritative on the original row).
    """

    row = session.get(ExportRun, export_run_id)
    if row is None or row.tenant_id != tenant_id:
        return None
    row.status = status
    if artifact_uri is not None:
        row.artifact_uri = artifact_uri
    if summary is not None:
        merged: dict[str, object] = {}
        if row.summary:
            try:
                existing = json.loads(row.summary)
                if isinstance(existing, dict):
                    merged.update(existing)
            except json.JSONDecodeError:
                pass
        merged.update(summary)
        row.summary = json.dumps(merged, sort_keys=True, ensure_ascii=False, default=str)
    if errors is not None:
        row.summary = _merge_errors(row.summary, errors)
    session.flush()
    return row


def _merge_errors(summary: str | None, errors: str) -> str:
    if summary:
        try:
            existing = json.loads(summary)
        except json.JSONDecodeError:
            existing = {}
    else:
        existing = {}
    if not isinstance(existing, dict):
        existing = {}
    existing["errors"] = errors
    return json.dumps(existing, sort_keys=True, ensure_ascii=False, default=str)


__all__ = [
    "ExportEnvelope",
    "MIME_XLSX",
    "PONTAJ_BLOCK_STARTS",
    "SCHEMA",
    "create_pending_export_run",
    "finalize_export_run",
    "record_export_run",
    "render_bulk_export",
    "render_pontaj_only_export",
    "render_store_export",
]
