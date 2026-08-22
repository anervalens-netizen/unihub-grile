from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from ugrile.core import database
from ugrile.domain.enums import MonthState
from ugrile.repositories.models import ManagerScope, Person, StoreAssignment
from ugrile.repositories.months import MonthRepository

MANAGER = {"X-Ugrile-Identity": "user_manager", "X-Ugrile-Tenant": "tenant_acme"}


def _open_scoped_month(faker_tenant) -> str:
    with database.session_scope() as session:
        month = MonthRepository(session).get_or_create(faker_tenant["tenant_id"], 2026, 8)
        month.state = MonthState.OPEN
        session.add(
            ManagerScope(
                tenant_id=faker_tenant["tenant_id"],
                user_id="user_manager",
                store_id=faker_tenant["store_id"],
                effective_from=date(2026, 8, 1),
                effective_to=date(2026, 8, 31),
            )
        )
        session.commit()
        return month.id


def _manager_rows(content: bytes) -> dict[str, tuple[object, object, int]]:
    workbook = load_workbook(BytesIO(content), data_only=False)
    sheet = workbook["user_manager"]
    return {
        str(sheet.cell(row, 1).value): (
            sheet.cell(row, 3).value,
            sheet.cell(row, 4).value,
            row,
        )
        for row in range(2, sheet.max_row + 1)
        if sheet.cell(row, 1).value
    }


def test_xlsx_catalog_uses_effective_home_not_mutable_current_home(client, faker_tenant):
    month_id = _open_scoped_month(faker_tenant)
    with database.session_scope() as session:
        alice = session.get(Person, faker_tenant["person_a_id"])
        carmen = session.get(Person, faker_tenant["person_c_id"])
        assert alice is not None and carmen is not None

        # Deliberately invert the mutable catalog values relative to dated history.
        alice.home_store_id = faker_tenant["other_store_id"]
        carmen.home_store_id = faker_tenant["store_id"]
        session.add_all(
            [
                StoreAssignment(
                    tenant_id=faker_tenant["tenant_id"],
                    person_id=alice.id,
                    store_id=faker_tenant["store_id"],
                    effective_from=date(2026, 8, 1),
                    effective_to=date(2026, 8, 31),
                ),
                StoreAssignment(
                    tenant_id=faker_tenant["tenant_id"],
                    person_id=carmen.id,
                    store_id=faker_tenant["other_store_id"],
                    effective_from=date(2026, 8, 1),
                    effective_to=date(2026, 8, 31),
                ),
            ]
        )
        session.commit()

    response = client.get(f"/months/{month_id}/schedule/template", headers=MANAGER)
    assert response.status_code == 200, response.text
    rows = _manager_rows(response.content)

    assert faker_tenant["person_a_id"] in rows
    assert rows[faker_tenant["person_a_id"]][0] == "s1"
    assert faker_tenant["person_c_id"] not in rows


def test_xlsx_transfer_or_history_gap_fails_closed_as_blocked_row(client, faker_tenant):
    month_id = _open_scoped_month(faker_tenant)
    with database.session_scope() as session:
        session.add_all(
            [
                StoreAssignment(
                    tenant_id=faker_tenant["tenant_id"],
                    person_id=faker_tenant["person_a_id"],
                    store_id=faker_tenant["store_id"],
                    effective_from=date(2026, 8, 1),
                    effective_to=date(2026, 8, 15),
                ),
                StoreAssignment(
                    tenant_id=faker_tenant["tenant_id"],
                    person_id=faker_tenant["person_a_id"],
                    store_id=faker_tenant["other_store_id"],
                    effective_from=date(2026, 8, 16),
                    effective_to=date(2026, 8, 31),
                ),
            ]
        )
        session.commit()

    response = client.get(f"/months/{month_id}/schedule/template", headers=MANAGER)
    assert response.status_code == 200, response.text
    workbook = load_workbook(BytesIO(response.content), data_only=False)
    sheet = workbook["user_manager"]
    alice_row = next(
        row
        for row in range(2, sheet.max_row + 1)
        if sheet.cell(row, 1).value == faker_tenant["person_a_id"]
    )

    assert sheet.cell(alice_row, 3).value in {None, ""}
    for column in range(4, 35):
        assert sheet.cell(alice_row, column).value == "BLOCAT"
        assert sheet.cell(alice_row, column).protection.locked is True


def test_xlsx_contract_detects_effective_home_history_drift(client, faker_tenant):
    month_id = _open_scoped_month(faker_tenant)
    with database.session_scope() as session:
        assignment = StoreAssignment(
            tenant_id=faker_tenant["tenant_id"],
            person_id=faker_tenant["person_a_id"],
            store_id=faker_tenant["store_id"],
            effective_from=date(2026, 8, 1),
            effective_to=date(2026, 8, 31),
        )
        session.add(assignment)
        session.commit()
        assignment_id = assignment.id

    template = client.get(f"/months/{month_id}/schedule/template", headers=MANAGER)
    assert template.status_code == 200, template.text

    # Change dated history without touching Month.revision. The import contract
    # must still become stale because person/date resource identity changed.
    with database.session_scope() as session:
        assignment = session.get(StoreAssignment, assignment_id)
        assert assignment is not None
        assignment.store_id = faker_tenant["other_store_id"]
        session.commit()

    preview = client.post(
        f"/months/{month_id}/schedule/preview",
        files={
            "file": (
                "schedule.xlsx",
                template.content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=MANAGER,
    )
    assert preview.status_code == 200, preview.text
    assert any(
        error.get("details", {}).get("code") == "CONTRACT_CHANGED"
        for error in preview.json()["errors"]
    )
