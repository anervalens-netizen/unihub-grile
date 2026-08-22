"""S2 calendar and XLSX schedule API tests."""

from __future__ import annotations

from datetime import date
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from ugrile.core import database
from ugrile.domain.enums import DayStatus, MonthState, RoleName, WorkingKind
from ugrile.repositories.models import (
    ManagerScope,
    Month,
    PersonDayAbsence,
    SiteDayAssignment,
    User,
)
from ugrile.repositories.months import MonthRepository

HEADERS = {"X-Ugrile-Identity": "user_admin", "X-Ugrile-Tenant": "tenant_acme"}
XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture()
def client(engine, faker_tenant):
    from ugrile.main import create_app

    with TestClient(create_app()) as test_client:
        yield test_client


def _open_month(faker_tenant):
    with database.session_scope() as session:
        month = MonthRepository(session).get_or_create(faker_tenant["tenant_id"], 2026, 8)
        month.state = MonthState.OPEN
        session.commit()
        return month.id


def _program_cell(
    client,
    month_id: str,
    *,
    expected_revision: int,
    person_id: str,
    business_date: str,
    status: str,
    store_id: str | None = None,
    working_kind: str | None = None,
):
    return client.post(
        f"/months/{month_id}/program/cell?expected_revision={expected_revision}",
        headers=HEADERS,
        json={
            "person_id": person_id,
            "business_date": business_date,
            "status": status,
            "store_id": store_id,
            "working_kind": working_kind,
        },
    )


def test_program_cells_derive_absence_and_revision(client, faker_tenant):
    month_id = _open_month(faker_tenant)
    first = _program_cell(
        client,
        month_id,
        expected_revision=0,
        person_id=faker_tenant["person_a_id"],
        business_date="2026-08-01",
        status="WORKING",
        store_id=faker_tenant["store_id"],
        working_kind="NORMAL",
    )
    assert first.status_code == 200, first.text
    second = _program_cell(
        client,
        month_id,
        expected_revision=1,
        person_id=faker_tenant["person_b_id"],
        business_date="2026-08-01",
        status="LEAVE",
    )
    assert second.status_code == 200, second.text
    third = _program_cell(
        client,
        month_id,
        expected_revision=2,
        person_id=faker_tenant["person_c_id"],
        business_date="2026-08-01",
        status="OFF",
    )
    assert third.status_code == 200, third.text
    body = third.json()
    assert body["revision"] == 3
    assert body["assignment_count"] == 1
    assert body["person_calendar_count"] == 3 * 31
    assert body["pontaj_count"] == 3 * 31

    with database.session_scope() as session:
        absences = session.query(PersonDayAbsence).filter_by(month_id=month_id).all()
        assert {(row.person_id, row.status) for row in absences} == {
            (faker_tenant["person_b_id"], DayStatus.LEAVE.value),
            (faker_tenant["person_c_id"], DayStatus.OFF.value),
        }
        assignments = session.query(SiteDayAssignment).filter_by(month_id=month_id).all()
        assert len(assignments) == 1

    stale = _program_cell(
        client,
        month_id,
        expected_revision=0,
        person_id=faker_tenant["person_a_id"],
        business_date="2026-08-02",
        status="OFF",
    )
    assert stale.status_code == 409


def test_xlsx_template_preview_and_atomic_apply(client, faker_tenant):
    month_id = _open_month(faker_tenant)
    template = client.get(f"/months/{month_id}/schedule/template", headers=HEADERS)
    assert template.status_code == 200, template.text
    assert template.headers["content-type"].startswith(XLSX)

    workbook = load_workbook(BytesIO(template.content))
    manager_sheet = workbook["user_admin"]
    manager_sheet["D2"] = "NORMAL - s1"
    edited = BytesIO()
    workbook.save(edited)
    edited_bytes = edited.getvalue()

    preview = client.post(
        f"/months/{month_id}/schedule/preview",
        files={"file": ("schedule.xlsx", edited_bytes, XLSX)},
        headers=HEADERS,
    )
    assert preview.status_code == 200, preview.text
    assert preview.json()["errors"] == []
    assert preview.json()["base_revision"] == 0

    applied = client.post(
        f"/months/{month_id}/schedule/apply",
        files={"file": ("schedule.xlsx", edited_bytes, XLSX)},
        headers=HEADERS,
    )
    assert applied.status_code == 200, applied.text
    assert applied.json()["revision"] == 1

    with database.session_scope() as session:
        assignments = session.query(SiteDayAssignment).filter_by(month_id=month_id).all()
        assert len(assignments) == 1
        assert assignments[0].person_id == faker_tenant["person_a_id"]

    refreshed = client.get(f"/months/{month_id}/schedule/template", headers=HEADERS)
    assert refreshed.status_code == 200
    refreshed_workbook = load_workbook(BytesIO(refreshed.content), data_only=True)
    assert refreshed_workbook["user_admin"]["D2"].value == "NORMAL - s1"

    stale_apply = client.post(
        f"/months/{month_id}/schedule/apply",
        files={"file": ("schedule.xlsx", edited_bytes, XLSX)},
        headers=HEADERS,
    )
    assert stale_apply.status_code == 409
    assert stale_apply.json()["code"] == "CONTRACT_CONSUMED"
    assert stale_apply.json()["details"]["code"] == "CONTRACT_CONSUMED"


def test_xlsx_preview_reports_coverage_and_kind_errors(client, faker_tenant):
    month_id = _open_month(faker_tenant)
    template = client.get(f"/months/{month_id}/schedule/template", headers=HEADERS)
    workbook = load_workbook(BytesIO(template.content))
    sheet = workbook["user_admin"]
    sheet["D2"] = "NORMAL - s1"
    sheet["D3"] = "NORMAL - s1"
    sheet["D4"] = "LIBER"
    edited = BytesIO()
    workbook.save(edited)

    conflict_preview = client.post(
        f"/months/{month_id}/schedule/preview",
        files={"file": ("schedule.xlsx", edited.getvalue(), XLSX)},
        headers=HEADERS,
    )
    assert conflict_preview.status_code == 200, conflict_preview.text
    conflict_codes = {error["code"] for error in conflict_preview.json()["errors"]}
    assert "COVERAGE_INVARIANT" in conflict_codes

    sheet["D2"] = "LIBER"
    sheet["D3"] = "LIBER"
    sheet["D4"] = "NORMAL - s1"
    invalid = BytesIO()
    workbook.save(invalid)
    invalid_preview = client.post(
        f"/months/{month_id}/schedule/preview",
        files={"file": ("schedule.xlsx", invalid.getvalue(), XLSX)},
        headers=HEADERS,
    )
    assert invalid_preview.status_code == 200, invalid_preview.text
    invalid_codes = {error["code"] for error in invalid_preview.json()["errors"]}
    assert "VALIDATION_ERROR" in invalid_codes


def test_xlsx_preview_reports_unknown_person_without_writes(client, faker_tenant):
    month_id = _open_month(faker_tenant)
    template = client.get(f"/months/{month_id}/schedule/template", headers=HEADERS)
    workbook = load_workbook(BytesIO(template.content))
    workbook["user_admin"]["A2"] = "person_acme_unknown"
    edited = BytesIO()
    workbook.save(edited)

    preview = client.post(
        f"/months/{month_id}/schedule/preview",
        files={"file": ("schedule.xlsx", edited.getvalue(), XLSX)},
        headers=HEADERS,
    )
    assert preview.status_code == 200
    assert any(error["code"] == "UNKNOWN_PERSON" for error in preview.json()["errors"])
    with database.session_scope() as session:
        assert session.query(SiteDayAssignment).filter_by(month_id=month_id).count() == 0


def test_manager_read_scope_filters_catalog_and_calendar(client, faker_tenant):
    month_id = _open_month(faker_tenant)
    manager_headers = {
        "X-Ugrile-Identity": "user_manager",
        "X-Ugrile-Tenant": "tenant_acme",
    }
    with database.session_scope() as session:
        manager = User(
            id="user_manager",
            tenant_id=faker_tenant["tenant_id"],
            email="manager@acme.example",
            display_name="Manager",
            role=RoleName.MANAGER.value,
        )
        session.add(manager)
        session.add(
            ManagerScope(
                tenant_id=faker_tenant["tenant_id"],
                user_id=manager.id,
                store_id=faker_tenant["store_id"],
                effective_from=date(2026, 8, 1),
            )
        )
        session.add_all(
            [
                SiteDayAssignment(
                    tenant_id=faker_tenant["tenant_id"],
                    month_id=month_id,
                    store_id=faker_tenant["store_id"],
                    person_id=faker_tenant["person_a_id"],
                    business_date=date(2026, 8, 1),
                    status=DayStatus.WORKING.value,
                    working_kind=WorkingKind.NORMAL.value,
                    revision=0,
                    source="TEST",
                ),
                SiteDayAssignment(
                    tenant_id=faker_tenant["tenant_id"],
                    month_id=month_id,
                    store_id=faker_tenant["other_store_id"],
                    person_id=faker_tenant["person_c_id"],
                    business_date=date(2026, 8, 1),
                    status=DayStatus.WORKING.value,
                    working_kind=WorkingKind.NORMAL.value,
                    revision=0,
                    source="TEST",
                ),
            ]
        )
        session.commit()

    assignments = client.get(f"/months/{month_id}/assignments", headers=manager_headers)
    assert assignments.status_code == 200, assignments.text
    assert [row["person_id"] for row in assignments.json()] == [faker_tenant["person_a_id"]]

    stores = client.get("/catalog/stores", headers=manager_headers)
    assert stores.status_code == 200, stores.text
    assert [row["id"] for row in stores.json()] == [faker_tenant["store_id"]]

    people = client.get("/catalog/people", headers=manager_headers)
    assert people.status_code == 200, people.text
    assert {row["id"] for row in people.json()} == {
        faker_tenant["person_a_id"],
        faker_tenant["person_b_id"],
    }


def _edit_cell(template_content: bytes, cell: str, value: str, sheet: str = "user_admin") -> bytes:
    workbook = load_workbook(BytesIO(template_content))
    workbook[sheet][cell] = value
    out = BytesIO()
    workbook.save(out)
    return out.getvalue()


def test_template_manifest_embeds_unique_contract_token(client, faker_tenant):
    month_id = _open_month(faker_tenant)
    first = client.get(f"/months/{month_id}/schedule/template", headers=HEADERS)
    second = client.get(f"/months/{month_id}/schedule/template", headers=HEADERS)
    assert first.status_code == 200 and second.status_code == 200
    wb1 = load_workbook(BytesIO(first.content), data_only=True)
    wb2 = load_workbook(BytesIO(second.content), data_only=True)
    token1 = wb1["Manifest"]["B5"].value
    token2 = wb2["Manifest"]["B5"].value
    assert token1 and token2
    assert token1 != token2
    assert wb1["Manifest"].protection.sheet is True
    assert wb1["Manifest"].sheet_state == "hidden"


def test_preview_never_consumes_the_contract(client, faker_tenant):
    month_id = _open_month(faker_tenant)
    template = client.get(f"/months/{month_id}/schedule/template", headers=HEADERS)
    edited = _edit_cell(template.content, "D2", "NORMAL - s1")
    for _ in range(2):
        preview = client.post(
            f"/months/{month_id}/schedule/preview",
            files={"file": ("schedule.xlsx", edited, XLSX)},
            headers=HEADERS,
        )
        assert preview.status_code == 200, preview.text
        assert preview.json()["errors"] == []
    applied = client.post(
        f"/months/{month_id}/schedule/apply",
        files={"file": ("schedule.xlsx", edited, XLSX)},
        headers=HEADERS,
    )
    assert applied.status_code == 200, applied.text
    assert applied.json()["revision"] == 1


def test_tampered_manifest_base_revision_is_rejected(client, faker_tenant):
    month_id = _open_month(faker_tenant)
    template = client.get(f"/months/{month_id}/schedule/template", headers=HEADERS)
    workbook = load_workbook(BytesIO(template.content))
    workbook["Manifest"]["B4"] = 99
    out = BytesIO()
    workbook.save(out)
    tampered = out.getvalue()

    preview = client.post(
        f"/months/{month_id}/schedule/preview",
        files={"file": ("schedule.xlsx", tampered, XLSX)},
        headers=HEADERS,
    )
    assert preview.status_code == 200, preview.text
    assert any(
        e.get("details", {}).get("code") == "MANIFEST_TAMPERED" for e in preview.json()["errors"]
    )

    applied = client.post(
        f"/months/{month_id}/schedule/apply",
        files={"file": ("schedule.xlsx", tampered, XLSX)},
        headers=HEADERS,
    )
    assert applied.status_code == 422
    assert applied.json()["details"]["code"] == "MANIFEST_TAMPERED"


def test_tampered_manifest_token_is_rejected(client, faker_tenant):
    month_id = _open_month(faker_tenant)
    template = client.get(f"/months/{month_id}/schedule/template", headers=HEADERS)
    workbook = load_workbook(BytesIO(template.content))
    workbook["Manifest"]["B5"] = "forged-token"
    out = BytesIO()
    workbook.save(out)
    tampered = out.getvalue()
    preview = client.post(
        f"/months/{month_id}/schedule/preview",
        files={"file": ("schedule.xlsx", tampered, XLSX)},
        headers=HEADERS,
    )
    assert any(
        e.get("details", {}).get("code") == "MANIFEST_TAMPERED" for e in preview.json()["errors"]
    )


def test_tampered_manifest_tenant_is_rejected(client, faker_tenant):
    month_id = _open_month(faker_tenant)
    template = client.get(f"/months/{month_id}/schedule/template", headers=HEADERS)
    workbook = load_workbook(BytesIO(template.content))
    workbook["Manifest"]["B2"] = "tenant_other"
    out = BytesIO()
    workbook.save(out)
    preview = client.post(
        f"/months/{month_id}/schedule/preview",
        files={"file": ("schedule.xlsx", out.getvalue(), XLSX)},
        headers=HEADERS,
    )
    assert any(e["code"] == "TENANT_MISMATCH" for e in preview.json()["errors"])
    assert any(
        e.get("details", {}).get("code") == "MANIFEST_TAMPERED" for e in preview.json()["errors"]
    )


def test_stale_contract_workbook_is_rejected(client, faker_tenant):
    """An unconsumed contract is stale when the calendar advanced elsewhere."""
    month_id = _open_month(faker_tenant)
    template = client.get(f"/months/{month_id}/schedule/template", headers=HEADERS)
    edited = _edit_cell(template.content, "D2", "NORMAL - s1")

    response = _program_cell(
        client,
        month_id,
        expected_revision=0,
        person_id=faker_tenant["person_a_id"],
        business_date="2026-08-01",
        status="WORKING",
        store_id=faker_tenant["store_id"],
        working_kind="NORMAL",
    )
    assert response.status_code == 200
    assert response.json()["revision"] == 1

    stale = client.post(
        f"/months/{month_id}/schedule/apply",
        files={"file": ("schedule.xlsx", edited, XLSX)},
        headers=HEADERS,
    )
    assert stale.status_code == 409
    assert stale.json()["code"] == "STALE_REVISION"
    with database.session_scope() as session:
        assert session.query(SiteDayAssignment).filter_by(month_id=month_id).count() == 1


def test_failed_apply_rolls_back_consumption_and_writes(client, faker_tenant):
    month_id = _open_month(faker_tenant)
    template = client.get(f"/months/{month_id}/schedule/template", headers=HEADERS)
    conflict = _edit_cell(_edit_cell(template.content, "D2", "NORMAL - s1"), "D3", "NORMAL - s1")
    failed = client.post(
        f"/months/{month_id}/schedule/apply",
        files={"file": ("schedule.xlsx", conflict, XLSX)},
        headers=HEADERS,
    )
    assert failed.status_code == 409
    assert failed.json()["code"] == "COVERAGE_INVARIANT"
    with database.session_scope() as session:
        assert session.query(SiteDayAssignment).filter_by(month_id=month_id).count() == 0
        assert session.query(Month).filter_by(id=month_id).one().revision == 0

    fixed = _edit_cell(conflict, "D3", "LIBER")
    applied = client.post(
        f"/months/{month_id}/schedule/apply",
        files={"file": ("schedule.xlsx", fixed, XLSX)},
        headers=HEADERS,
    )
    assert applied.status_code == 200, applied.text
    assert applied.json()["revision"] == 1


def test_partial_month_manager_scope_blocks_out_of_scope_days(client, faker_tenant):
    month_id = _open_month(faker_tenant)
    for day in range(1, 6):
        initial = _program_cell(
            client,
            month_id,
            expected_revision=day - 1,
            person_id=faker_tenant["person_a_id"],
            business_date=f"2026-08-{day:02d}",
            status="WORKING",
            store_id=faker_tenant["store_id"],
            working_kind="NORMAL",
        )
        assert initial.status_code == 200, initial.text
    assert initial.json()["revision"] == 5

    manager_headers = {
        "X-Ugrile-Identity": "user_manager",
        "X-Ugrile-Tenant": "tenant_acme",
    }
    with database.session_scope() as session:
        session.add(
            User(
                id="user_manager",
                tenant_id=faker_tenant["tenant_id"],
                email="manager@acme.example",
                display_name="Manager",
                role=RoleName.MANAGER.value,
            )
        )
        session.add(
            ManagerScope(
                tenant_id=faker_tenant["tenant_id"],
                user_id="user_manager",
                store_id=faker_tenant["store_id"],
                effective_from=date(2026, 8, 1),
                effective_to=date(2026, 8, 15),
            )
        )
        session.commit()

    template = client.get(f"/months/{month_id}/schedule/template", headers=manager_headers)
    assert template.status_code == 200, template.text
    wb = load_workbook(BytesIO(template.content), data_only=True)
    sheet = wb["user_manager"]
    assert sheet["D2"].value == "NORMAL - s1"
    assert sheet["I2"].value == "LIBER"
    assert sheet["S2"].value == "BLOCAT"
    assert sheet["S2"].protection.locked is True
    assert sheet["I2"].protection.locked is False
    assert sheet["AH2"].value == "BLOCAT"

    edited = _edit_cell(template.content, "E2", "LIBER", sheet="user_manager")
    applied = client.post(
        f"/months/{month_id}/schedule/apply",
        files={"file": ("schedule.xlsx", edited, XLSX)},
        headers=manager_headers,
    )
    assert applied.status_code == 200, applied.text
    assert applied.json()["revision"] == 6
    with database.session_scope() as session:
        working = {
            row.business_date.day
            for row in session.query(SiteDayAssignment).filter_by(month_id=month_id)
            if row.status == DayStatus.WORKING.value
        }
    assert working == {1, 3, 4, 5}

    pristine = client.get(f"/months/{month_id}/schedule/template", headers=manager_headers)
    reapply = client.post(
        f"/months/{month_id}/schedule/apply",
        files={"file": ("schedule.xlsx", pristine.content, XLSX)},
        headers=manager_headers,
    )
    assert reapply.status_code == 200, reapply.text
    with database.session_scope() as session:
        working = {
            row.business_date.day
            for row in session.query(SiteDayAssignment).filter_by(month_id=month_id)
            if row.status == DayStatus.WORKING.value
        }
    assert working == {1, 3, 4, 5}


def test_tampered_blocked_cell_is_rejected_without_writes(client, faker_tenant):
    month_id = _open_month(faker_tenant)
    manager_headers = {
        "X-Ugrile-Identity": "user_manager",
        "X-Ugrile-Tenant": "tenant_acme",
    }
    with database.session_scope() as session:
        session.add(
            User(
                id="user_manager",
                tenant_id=faker_tenant["tenant_id"],
                email="manager@acme.example",
                display_name="Manager",
                role=RoleName.MANAGER.value,
            )
        )
        session.add(
            ManagerScope(
                tenant_id=faker_tenant["tenant_id"],
                user_id="user_manager",
                store_id=faker_tenant["store_id"],
                effective_from=date(2026, 8, 1),
                effective_to=date(2026, 8, 15),
            )
        )
        session.commit()

    template = client.get(f"/months/{month_id}/schedule/template", headers=manager_headers)
    wb = load_workbook(BytesIO(template.content))
    sheet = wb["user_manager"]
    assert sheet["S2"].value == "BLOCAT"
    sheet["S2"] = "NORMAL - s1"
    out = BytesIO()
    wb.save(out)
    tampered = out.getvalue()

    preview = client.post(
        f"/months/{month_id}/schedule/preview",
        files={"file": ("schedule.xlsx", tampered, XLSX)},
        headers=manager_headers,
    )
    assert preview.status_code == 200, preview.text
    assert any(e["code"] == "FORBIDDEN" for e in preview.json()["errors"])

    applied = client.post(
        f"/months/{month_id}/schedule/apply",
        files={"file": ("schedule.xlsx", tampered, XLSX)},
        headers=manager_headers,
    )
    assert applied.status_code == 403
    with database.session_scope() as session:
        assert session.query(SiteDayAssignment).filter_by(month_id=month_id).count() == 0
        assert session.query(Month).filter_by(id=month_id).one().revision == 0


def test_pontaj_endpoint_reads_persisted_projection(client, faker_tenant):
    month_id = _open_month(faker_tenant)
    working_response = _program_cell(
        client,
        month_id,
        expected_revision=0,
        person_id=faker_tenant["person_a_id"],
        business_date="2026-08-01",
        status="WORKING",
        store_id=faker_tenant["store_id"],
        working_kind="NORMAL",
    )
    assert working_response.status_code == 200, working_response.text
    leave_response = _program_cell(
        client,
        month_id,
        expected_revision=1,
        person_id=faker_tenant["person_b_id"],
        business_date="2026-08-02",
        status="LEAVE",
    )
    assert leave_response.status_code == 200, leave_response.text
    assert leave_response.json()["revision"] == 2

    pontaj = client.get(f"/months/{month_id}/pontaj", headers=HEADERS)
    assert pontaj.status_code == 200, pontaj.text
    body = pontaj.json()
    assert body["revision"] == 2
    assert len(body["rows"]) == 3 * 31
    working = next(
        r
        for r in body["rows"]
        if r["person_id"] == faker_tenant["person_a_id"] and r["business_date"] == "2026-08-01"
    )
    assert working["status"] == "WORKING"
    assert working["hours"] == "11.00"
    assert working["start_time"] == "10:00:00"
    assert working["end_time"] == "22:00:00"
    assert working["pause_minutes"] == 60
    leave = next(
        r
        for r in body["rows"]
        if r["person_id"] == faker_tenant["person_b_id"] and r["business_date"] == "2026-08-02"
    )
    assert leave["status"] == "LEAVE"
    assert leave["hours"] == "0.00"
    assert leave["start_time"] is None

    totals = {t["person_id"]: t for t in body["totals"]}
    assert totals[faker_tenant["person_a_id"]]["working_days"] == 1
    assert totals[faker_tenant["person_a_id"]]["total_hours"] == "11.00"
    assert totals[faker_tenant["person_b_id"]]["leave_days"] == 1
    assert totals[faker_tenant["person_c_id"]]["off_days"] == 31


def test_pontaj_endpoint_filters_manager_scope(client, faker_tenant):
    month_id = _open_month(faker_tenant)
    response = _program_cell(
        client,
        month_id,
        expected_revision=0,
        person_id=faker_tenant["person_a_id"],
        business_date="2026-08-01",
        status="WORKING",
        store_id=faker_tenant["store_id"],
        working_kind="NORMAL",
    )
    assert response.status_code == 200, response.text
    manager_headers = {
        "X-Ugrile-Identity": "user_manager",
        "X-Ugrile-Tenant": "tenant_acme",
    }
    with database.session_scope() as session:
        session.add(
            User(
                id="user_manager",
                tenant_id=faker_tenant["tenant_id"],
                email="manager@acme.example",
                display_name="Manager",
                role=RoleName.MANAGER.value,
            )
        )
        session.add(
            ManagerScope(
                tenant_id=faker_tenant["tenant_id"],
                user_id="user_manager",
                store_id=faker_tenant["store_id"],
                effective_from=date(2026, 8, 1),
                effective_to=date(2026, 8, 15),
            )
        )
        session.commit()

    pontaj = client.get(f"/months/{month_id}/pontaj", headers=manager_headers)
    assert pontaj.status_code == 200, pontaj.text
    body = pontaj.json()
    assert len(body["rows"]) == 2 * 15
    assert all(r["business_date"] <= "2026-08-15" for r in body["rows"])
    assert all(r["person_id"] != faker_tenant["person_c_id"] for r in body["rows"])
    totals = {t["person_id"]: t for t in body["totals"]}
    assert totals[faker_tenant["person_a_id"]]["total_hours"] == "11.00"
