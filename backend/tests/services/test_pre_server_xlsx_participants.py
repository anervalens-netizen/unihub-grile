from __future__ import annotations

import io
from datetime import date, time
from decimal import Decimal

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import select

from ugrile.domain.errors import ValidationError
from ugrile.repositories.models import Person, PontajProjection, SiteDayAssignment
from ugrile.repositories.months import MonthRepository
from ugrile.services.revisioned_xlsx_export import render_store_export_at_revision
from ugrile.services.xlsx_export import PONTAJ_BLOCK_STARTS, _write_pontaj_tab


def test_revisioned_store_pontaj_renders_third_historical_participant(
    session, faker_tenant
) -> None:
    tenant_id = faker_tenant["tenant_id"]
    month = MonthRepository(session).get_or_create(tenant_id, 2026, 8)
    people = list(
        session.execute(
            select(Person)
            .where(
                Person.tenant_id == tenant_id,
                Person.id.in_(
                    [
                        faker_tenant["person_a_id"],
                        faker_tenant["person_b_id"],
                        faker_tenant["person_c_id"],
                    ]
                ),
            )
            .order_by(Person.internal_code)
        ).scalars()
    )
    assert len(people) == 3
    people[2].is_active = False

    for index, person in enumerate(people, start=1):
        business_date = date(2026, 8, index)
        session.add(
            SiteDayAssignment(
                tenant_id=tenant_id,
                month_id=month.id,
                store_id=faker_tenant["store_id"],
                person_id=person.id,
                business_date=business_date,
                status="WORKING",
                working_kind=(
                    "EXTRA_OTHER"
                    if person.id == faker_tenant["person_c_id"]
                    else "NORMAL"
                ),
                revision=1,
                source="PRE_SERVER_TEST",
            )
        )
        session.add(
            PontajProjection(
                tenant_id=tenant_id,
                month_id=month.id,
                person_id=person.id,
                business_date=business_date,
                revision=1,
                status="WORKING",
                start_time=time(10, 0),
                end_time=time(22, 0),
                pause_minutes=60,
                hours=Decimal("11.00"),
            )
        )
    session.commit()

    envelope = render_store_export_at_revision(
        session,
        tenant_id=tenant_id,
        month=month,
        store_id=faker_tenant["store_id"],
        revision=1,
    )
    workbook = load_workbook(io.BytesIO(envelope.bytes_))
    pontaj = workbook["Pontaj"]
    assert [pontaj.cell(row=row, column=1).value for row in (8, 11, 14)] == [
        person.display_name for person in people
    ]
    assert pontaj.cell(row=14, column=5).value == "11.00"


def test_pontaj_fails_closed_above_layout_capacity(faker_tenant) -> None:
    assert len(PONTAJ_BLOCK_STARTS) == 8
    people = {
        f"person_capacity_{index}": Person(
            id=f"person_capacity_{index}",
            tenant_id=faker_tenant["tenant_id"],
            internal_code=f"cap{index:02d}",
            display_name=f"Capacity {index}",
            home_store_id=faker_tenant["store_id"],
        )
        for index in range(9)
    }
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None

    with pytest.raises(ValidationError) as exc_info:
        _write_pontaj_tab(
            sheet,
            pontaj_rows=[],
            assignments_by_person_day={},
            persons_by_id=people,
            month_year=2026,
            month_month=8,
        )

    assert exc_info.value.details == {
        "code": "PONTAJ_LAYOUT_CAPACITY_EXCEEDED",
        "capacity": 8,
        "participants": 9,
    }
