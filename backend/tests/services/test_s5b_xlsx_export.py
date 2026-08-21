"""S5b export service tests (AC-14)."""

from __future__ import annotations

import hashlib
import io
import json
import zipfile
from datetime import date
from decimal import Decimal

from openpyxl import load_workbook
from sqlalchemy import select

from ugrile.connectors.fixtures import FIXTURE_GENERATION, default_fixture
from ugrile.connectors.ingest import FixtureConnector
from ugrile.domain.enums import DayStatus, MonthState, WorkingKind
from ugrile.repositories.models import SalesPersonDayProjection
from ugrile.repositories.months import MonthRepository
from ugrile.repositories.salary import SalaryRepository
from ugrile.services.calendar import CalendarChange, CalendarService
from ugrile.services.grid import GridService
from ugrile.services.xlsx_export import (
    render_bulk_export,
    render_pontaj_only_export,
    render_store_export,
)


def _seed_full_month(session, faker_tenant, store_id: str, person_id: str) -> None:
    FixtureConnector(session).apply(default_fixture())
    month = MonthRepository(session).get_or_create(faker_tenant["tenant_id"], 2026, 8)
    month.state = MonthState.OPEN
    SalaryRepository(session).upsert_window(
        tenant_id=faker_tenant["tenant_id"],
        person_id=person_id,
        effective_from=date(2026, 1, 1),
        effective_to=None,
        salary=Decimal("2600"),
        tickets=Decimal("480"),
        flip=Decimal("0"),
        source="HR_MASTER",
    )
    for day in range(1, 32):
        CalendarService(session).apply(
            month=month,
            tenant_id=faker_tenant["tenant_id"],
            expected_revision=month.revision,
            changes=[
                CalendarChange(
                    person_id,
                    date(2026, 8, day),
                    store_id,
                    DayStatus.WORKING,
                    WorkingKind.NORMAL,
                )
            ],
        )
        if (
            not session.execute(
                select(SalesPersonDayProjection).where(
                    SalesPersonDayProjection.tenant_id == faker_tenant["tenant_id"],
                    SalesPersonDayProjection.business_date == date(2026, 8, day),
                )
            )
            .scalars()
            .first()
        ):
            session.add(
                SalesPersonDayProjection(
                    tenant_id=faker_tenant["tenant_id"],
                    month_id=month.id,
                    person_id=person_id,
                    store_id=store_id,
                    business_date=date(2026, 8, day),
                    revision=month.revision + 1,
                    amount=Decimal("12500"),
                    currency="RON",
                    generation=FIXTURE_GENERATION,
                    working_kind="NORMAL",
                )
            )
    GridService(session).compute_and_persist(tenant_id=faker_tenant["tenant_id"], month=month)
    session.commit()


def test_store_export_renders_grila_and_pontaj_tabs(session, faker_tenant):
    _seed_full_month(
        session,
        faker_tenant,
        store_id=faker_tenant["store_id"],
        person_id=faker_tenant["person_a_id"],
    )
    month = MonthRepository(session).get_or_create(faker_tenant["tenant_id"], 2026, 8)
    envelope = render_store_export(
        session,
        tenant_id=faker_tenant["tenant_id"],
        month=month,
        store_id=faker_tenant["store_id"],
    )
    assert envelope.filename.endswith(".xlsx")
    assert envelope.checksum == hashlib.sha256(envelope.bytes_).hexdigest()
    workbook = load_workbook(io.BytesIO(envelope.bytes_))
    assert set(workbook.sheetnames) == {"Grila", "Pontaj"}
    grila = workbook["Grila"]
    assert grila["A1"].value.startswith("Magazin:")
    pontaj = workbook["Pontaj"]
    assert pontaj.cell(row=1, column=1).value == "Persoana"
    # day 1..31 in columns C..AG (3..33) + total AH at column 34
    assert pontaj.cell(row=1, column=3).value == 1
    assert pontaj.cell(row=1, column=33).value == 31
    assert pontaj.cell(row=1, column=34).value == "Total ore (AH)"


def test_pontaj_uses_standard_c8_ag31_block_layout(session, faker_tenant):
    """Per docs/MOBIUP_RULE_PACK.md §7, the standard Pontaj tab uses
    block rows 8 and 11 (the 3-row block: net hours / interval / pause)
    with column C..AG = day 1..31 and column AH = total."""
    from ugrile.domain.enums import DayStatus, WorkingKind
    from ugrile.services.calendar import CalendarChange, CalendarService

    _seed_full_month(
        session,
        faker_tenant,
        store_id=faker_tenant["store_id"],
        person_id=faker_tenant["person_a_id"],
    )
    # Seed WORKING for every day of the month so the total reflects 31 days x 11 h.
    month = MonthRepository(session).get_or_create(faker_tenant["tenant_id"], 2026, 8)
    for day in range(1, 32):
        CalendarService(session).apply(
            month=month,
            tenant_id=faker_tenant["tenant_id"],
            expected_revision=month.revision,
            changes=[
                CalendarChange(
                    faker_tenant["person_a_id"],
                    __import__("datetime").date(2026, 8, day),
                    faker_tenant["store_id"],
                    DayStatus.WORKING,
                    WorkingKind.NORMAL,
                )
            ],
        )
    session.commit()
    envelope = render_store_export(
        session,
        tenant_id=faker_tenant["tenant_id"],
        month=month,
        store_id=faker_tenant["store_id"],
    )
    workbook = load_workbook(io.BytesIO(envelope.bytes_))
    pontaj = workbook["Pontaj"]
    # Single-agent store uses block row 8 (rows 8/9/10 = net/interval/pause).
    assert pontaj.cell(row=8, column=1).value
    assert "Interval:" in str(pontaj.cell(row=9, column=1).value)
    assert "Pauza:" in str(pontaj.cell(row=10, column=1).value)
    # day 1 in column C (3) must carry 11.00 hours.
    assert pontaj.cell(row=8, column=3).value == "11.00"
    # Total in column AH (34) for the single-agent block at row 8: 31 x 11 = 341.
    total_cell = pontaj.cell(row=8, column=34).value
    assert float(total_cell.replace(",", ".")) == 31 * 11


def test_store_export_includes_only_store_assignments(session, faker_tenant):
    """Strict per-store filter: an export for store_a must not include
    Pontaj rows for store_b's persons."""
    _seed_full_month(
        session,
        faker_tenant,
        store_id=faker_tenant["store_id"],
        person_id=faker_tenant["person_a_id"],
    )
    # Add assignments + Pontaj rows on other_store via Carmen.
    month = MonthRepository(session).get_or_create(faker_tenant["tenant_id"], 2026, 8)
    for day in range(1, 32):
        CalendarService(session).apply(
            month=month,
            tenant_id=faker_tenant["tenant_id"],
            expected_revision=month.revision,
            changes=[
                CalendarChange(
                    faker_tenant["person_c_id"],
                    date(2026, 8, day),
                    faker_tenant["other_store_id"],
                    DayStatus.WORKING,
                    WorkingKind.NORMAL,
                )
            ],
        )
    session.commit()
    envelope = render_store_export(
        session,
        tenant_id=faker_tenant["tenant_id"],
        month=month,
        store_id=faker_tenant["store_id"],
    )
    workbook = load_workbook(io.BytesIO(envelope.bytes_))
    pontaj = workbook["Pontaj"]
    # Person c (Carmen, from other_store) must NOT appear in store_a's Pontaj.
    rendered = [pontaj.cell(row=r, column=1).value for r in range(2, 14)]
    assert not any("Carmen" in str(v) for v in rendered), rendered


def test_pontaj_only_export_spans_every_person(session, faker_tenant):
    _seed_full_month(
        session,
        faker_tenant,
        store_id=faker_tenant["store_id"],
        person_id=faker_tenant["person_a_id"],
    )
    month = MonthRepository(session).get_or_create(faker_tenant["tenant_id"], 2026, 8)
    envelope = render_pontaj_only_export(session, tenant_id=faker_tenant["tenant_id"], month=month)
    workbook = load_workbook(io.BytesIO(envelope.bytes_))
    assert workbook.active is not None
    ws = workbook.active
    assert ws.title == "Pontaj"


def test_bulk_export_zips_per_store_with_manifest(session, faker_tenant):
    _seed_full_month(
        session,
        faker_tenant,
        store_id=faker_tenant["store_id"],
        person_id=faker_tenant["person_a_id"],
    )
    month = MonthRepository(session).get_or_create(faker_tenant["tenant_id"], 2026, 8)
    envelope = render_bulk_export(session, tenant_id=faker_tenant["tenant_id"], month=month)
    with zipfile.ZipFile(io.BytesIO(envelope.bytes_)) as zf:
        names = zf.namelist()
        manifest_bytes = zf.read("manifest.json")
    manifest = json.loads(manifest_bytes)
    assert "manifest.json" in names
    assert manifest["schema"] == "UGRILE-S5-XLSX-V2"
    # Frozen AC-14 contract: manifest MUST carry the source generation
    # and the rule pack version that produced the workbooks. Source is
    # the connector generation (FIXTURE_V1 at S5) and the canonical V1
    # rule pack ("mobiup-v1-compat").
    assert manifest["generation"] == "FIXTURE_V1"
    assert manifest["rule_pack_version"] == "mobiup-v1-compat"
    assert manifest["store_count"] >= 1
    # Each entry's checksum must match the bytes persisted in the ZIP
    # (proves the manifest is bound to the actual workbook, not an
    # independent commissioning field).
    with zipfile.ZipFile(io.BytesIO(envelope.bytes_)) as zf:
        manifest_checksums = {
            entry["filename"]: entry["checksum_sha256"] for entry in manifest["entries"]
        }
        for filename, expected_checksum in manifest_checksums.items():
            assert (
                hashlib.sha256(zf.read(filename)).hexdigest() == expected_checksum
            ), f"checksum mismatch for {filename}"
