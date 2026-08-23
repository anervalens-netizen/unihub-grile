"""XLSX-001..003 production artifact proofs for the durable renderer."""

from __future__ import annotations

import hashlib
import io
import json
import time
import zipfile
from datetime import date

from openpyxl import load_workbook

from ugrile.domain.enums import DayStatus, MonthState, WorkingKind
from ugrile.repositories.months import MonthRepository
from ugrile.services.calendar import CalendarChange, CalendarService
from ugrile.services.revisioned_xlsx_export import (
    render_bulk_export_at_revision,
    render_pontaj_only_export_at_revision,
    render_store_export_at_revision,
)


def _seed_revision_one(session, tenant: dict[str, str]):
    month = MonthRepository(session).get_or_create(tenant["tenant_id"], 2026, 8)
    month.state = MonthState.OPEN
    result = CalendarService(session).apply(
        month=month,
        tenant_id=tenant["tenant_id"],
        expected_revision=month.revision,
        changes=[
            CalendarChange(
                tenant["person_a_id"],
                date(2026, 8, 1),
                tenant["store_id"],
                DayStatus.WORKING,
                WorkingKind.NORMAL,
            )
        ],
    )
    session.flush()
    assert result.revision == 1
    return month


def _assert_no_external_dependencies(payload: bytes) -> None:
    with zipfile.ZipFile(io.BytesIO(payload), mode="r") as archive:
        names = archive.namelist()
        assert not any(name.startswith("xl/externalLinks/") for name in names)
        assert "xl/externalLinks/_rels/externalLink1.xml.rels" not in names

    workbook = load_workbook(io.BytesIO(payload), data_only=False, keep_links=True)
    assert not getattr(workbook, "_external_links", [])
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                assert cell.data_type != "f", (
                    f"unexpected formula dependency in {sheet.title}!{cell.coordinate}: "
                    f"{cell.value!r}"
                )


def test_revision_pinned_store_export_is_byte_deterministic_and_parseable(
    session,
    faker_tenant,
) -> None:
    month = _seed_revision_one(session, faker_tenant)

    first = render_store_export_at_revision(
        session,
        tenant_id=faker_tenant["tenant_id"],
        month=month,
        store_id=faker_tenant["store_id"],
        revision=1,
    )
    # OpenPyXL normally changes core document timestamps between saves. Crossing
    # a wall-clock second proves the artifact canonicalizer removes that source
    # of checksum drift rather than accidentally passing within one second.
    time.sleep(1.05)
    second = render_store_export_at_revision(
        session,
        tenant_id=faker_tenant["tenant_id"],
        month=month,
        store_id=faker_tenant["store_id"],
        revision=1,
    )

    assert first.bytes_ == second.bytes_
    assert first.checksum == second.checksum == hashlib.sha256(first.bytes_).hexdigest()
    assert first.summary["revision"] == 1

    workbook = load_workbook(io.BytesIO(first.bytes_), data_only=False, keep_links=True)
    assert workbook.sheetnames == ["Grila", "Pontaj"]
    assert f"Revizie: 1" in str(workbook["Grila"]["A4"].value)
    assert workbook["Pontaj"]["A1"].value == "Persoana"
    assert workbook["Pontaj"]["AH1"].value == "Total ore (AH)"
    assert workbook["Pontaj"]["A8"].value == faker_tenant["person_a_name"]
    assert workbook["Pontaj"]["C8"].value == "11.00"
    _assert_no_external_dependencies(first.bytes_)


def test_revision_pinned_bulk_manifest_matches_embedded_workbook_checksum(
    session,
    faker_tenant,
) -> None:
    month = _seed_revision_one(session, faker_tenant)
    envelope = render_bulk_export_at_revision(
        session,
        tenant_id=faker_tenant["tenant_id"],
        month=month,
        revision=1,
        store_ids=[faker_tenant["store_id"]],
    )

    with zipfile.ZipFile(io.BytesIO(envelope.bytes_), mode="r") as archive:
        manifest = json.loads(archive.read("manifest.json"))
        assert manifest["revision"] == 1
        assert manifest["store_count"] == 1
        assert len(manifest["entries"]) == 1
        entry = manifest["entries"][0]
        assert entry["store_id"] == faker_tenant["store_id"]
        workbook_bytes = archive.read(entry["filename"])
        assert entry["size_bytes"] == len(workbook_bytes)
        assert entry["checksum_sha256"] == hashlib.sha256(workbook_bytes).hexdigest()
        _assert_no_external_dependencies(workbook_bytes)

    repeated = render_bulk_export_at_revision(
        session,
        tenant_id=faker_tenant["tenant_id"],
        month=month,
        revision=1,
        store_ids=[faker_tenant["store_id"]],
    )
    assert repeated.bytes_ == envelope.bytes_
    assert repeated.checksum == envelope.checksum


def test_revision_pinned_pontaj_only_is_scoped_parseable_and_dependency_free(
    session,
    faker_tenant,
) -> None:
    month = _seed_revision_one(session, faker_tenant)
    envelope = render_pontaj_only_export_at_revision(
        session,
        tenant_id=faker_tenant["tenant_id"],
        month=month,
        revision=1,
        store_ids=[faker_tenant["store_id"]],
    )

    assert envelope.summary["stores_included"] == [faker_tenant["store_id"]]
    assert envelope.summary["revision"] == 1
    workbook = load_workbook(io.BytesIO(envelope.bytes_), data_only=False, keep_links=True)
    assert workbook.sheetnames == ["Pontaj"]
    assert workbook["Pontaj"]["A8"].value == faker_tenant["person_a_name"]
    assert workbook["Pontaj"]["C8"].value == "11.00"
    _assert_no_external_dependencies(envelope.bytes_)
