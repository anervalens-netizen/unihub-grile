from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from ugrile.domain.enums import DayStatus, WorkingKind
from ugrile.services.calendar import CalendarChange
from ugrile.services.schedule_xlsx import SCHEMA, build_template, parse_schedule

TOKEN = "test-contract-token-abc123"


def _template(**overrides):
    kwargs = dict(
        tenant_id="tenant_acme",
        month_id="month_acme_2026_08",
        year=2026,
        month=8,
        base_revision=4,
        contract_token=TOKEN,
        people=[
            {
                "person_id": "person_acme_a",
                "display_name": "Alice",
                "home_store_code": "s1",
                "manager_code": "m1",
            }
        ],
        stores={"s1": "store_acme_s1"},
    )
    kwargs.update(overrides)
    return build_template(**kwargs)


def test_template_parsed_round_trip_and_manifest():
    data = _template(
        calendar=[
            CalendarChange(
                "person_acme_a",
                date(2026, 8, 1),
                "store_acme_s1",
                DayStatus.WORKING,
                WorkingKind.NORMAL,
            )
        ],
    )
    wb = load_workbook(BytesIO(data), data_only=True)
    assert {"Instrucțiuni", "Manifest", "_Lists", "m1"}.issubset(wb.sheetnames)
    assert wb["Manifest"]["B1"].value == SCHEMA
    assert wb["Manifest"]["B5"].value == TOKEN  # contract_token row
    assert wb["m1"].column_dimensions["A"].hidden is True
    assert wb["m1"].protection.sheet is True
    assert wb["m1"]["D2"].protection.locked is False
    assert wb["_Lists"].sheet_state == "hidden"
    assert wb["_Lists"].protection.sheet is True
    assert wb["m1"]["D2"].value == "NORMAL - s1"
    parsed = parse_schedule(
        data,
        expected_tenant_id="tenant_acme",
        expected_month_id="month_acme_2026_08",
        year=2026,
        month=8,
        stores={"s1": "store_acme_s1"},
    )
    assert parsed.errors == []
    assert parsed.base_revision == 4
    assert parsed.contract_token == TOKEN
    assert len(parsed.changes) == 31
    assert parsed.changes[0].status == DayStatus.WORKING


def test_xlsx_unknown_and_malformed_cells_are_reported():
    data = _template()
    wb = load_workbook(BytesIO(data))
    ws = wb["m1"]
    ws["D2"] = "NORMAL - does-not-exist"
    ws["E2"] = "BROKEN"
    out = BytesIO()
    wb.save(out)
    parsed = parse_schedule(
        out.getvalue(),
        expected_tenant_id="tenant_acme",
        expected_month_id="month_acme_2026_08",
        year=2026,
        month=8,
        stores={"s1": "store_acme_s1"},
    )
    assert {e["code"] for e in parsed.errors} == {"UNKNOWN_STORE", "MALFORMED_CELL"}


def test_template_dropdowns_are_person_specific():
    data = _template(stores={"s1": "store_acme_s1", "s2": "store_acme_s2"})
    workbook = load_workbook(BytesIO(data), data_only=True)
    options = [workbook["_Lists"].cell(row, 1).value for row in range(1, 6)]
    assert options == [
        "LIBER",
        "CONCEDIU",
        "NORMAL - s1",
        "SUPLIMENTAR ACASĂ - s1",
        "SUPLIMENTAR - s2",
    ]


def test_blocked_days_render_blocat_and_parse_as_no_change():
    """A person outside the date-specific scope gets locked BLOCAT cells."""
    data = _template(
        people=[
            {
                "person_id": "person_acme_a",
                "display_name": "Alice",
                "home_store_code": "s1",
                "manager_code": "m1",
            },
            {
                "person_id": "person_acme_c",
                "display_name": "Carmen",
                "home_store_code": "s2",
                "manager_code": "m1",
            },
        ],
        stores={"s1": "store_acme_s1", "s2": "store_acme_s2"},
        allowed_store_ids_by_date={date(2026, 8, d): {"store_acme_s1"} for d in range(1, 16)},
    )
    wb = load_workbook(BytesIO(data), data_only=True)
    # Alice is in scope 1..15 and blocked afterwards; Carmen is always blocked.
    # Column for day d is d+3: day 1 -> D, day 16 -> S.
    assert wb["m1"]["D2"].value == "LIBER"
    assert wb["m1"]["S2"].value == "BLOCAT"  # day 16
    assert wb["m1"]["D3"].value == "BLOCAT"  # Carmen, day 1
    parsed = parse_schedule(
        data,
        expected_tenant_id="tenant_acme",
        expected_month_id="month_acme_2026_08",
        year=2026,
        month=8,
        stores={"s1": "store_acme_s1", "s2": "store_acme_s2"},
    )
    assert parsed.errors == []
    # 15 in-scope days for Alice -> 15 changes; Carmen's blocked days produce none.
    assert len(parsed.changes) == 15
    assert all(change.person_id == "person_acme_a" for change in parsed.changes)
    assert {change.business_date for change in parsed.changes} == {
        date(2026, 8, d) for d in range(1, 16)
    }


def test_blocked_cell_tampering_is_rejected_instead_of_silent_off():
    data = _template(
        people=[
            {
                "person_id": "person_acme_c",
                "display_name": "Carmen",
                "home_store_code": "s2",
                "manager_code": "m1",
            }
        ],
        stores={"s1": "store_acme_s1", "s2": "store_acme_s2"},
        allowed_store_ids_by_date={date(2026, 8, 1): {"store_acme_s1"}},
    )
    wb = load_workbook(BytesIO(data))
    wb["m1"]["D2"] = "NORMAL - s1"  # tamper a BLOCAT cell
    out = BytesIO()
    wb.save(out)
    parsed = parse_schedule(
        out.getvalue(),
        expected_tenant_id="tenant_acme",
        expected_month_id="month_acme_2026_08",
        year=2026,
        month=8,
        stores={"s1": "store_acme_s1", "s2": "store_acme_s2"},
    )
    assert parsed.errors == []  # syntax-level parse is fine
    # The tampered change is present but the server-side scope rejects it at
    # apply time (tested at service/API level); the parser must not invent an
    # OFF change for untouched blocked cells.
    assert any(
        c.person_id == "person_acme_c" and c.business_date == date(2026, 8, 1)
        for c in parsed.changes
    )


def test_formula_cells_are_rejected():
    data = _template()
    wb = load_workbook(BytesIO(data))
    wb["m1"]["D2"] = '=IF(A1="x","LIBER","CONCEDIU")'
    out = BytesIO()
    wb.save(out)
    parsed = parse_schedule(
        out.getvalue(),
        expected_tenant_id="tenant_acme",
        expected_month_id="month_acme_2026_08",
        year=2026,
        month=8,
        stores={"s1": "store_acme_s1"},
    )
    assert {"FORMULA_CELL"} <= {e["code"] for e in parsed.errors}


def test_merged_day_cells_are_rejected():
    data = _template()
    wb = load_workbook(BytesIO(data))
    wb["m1"].merge_cells("D2:E2")
    wb["m1"]["D2"] = "LIBER"
    out = BytesIO()
    wb.save(out)
    parsed = parse_schedule(
        out.getvalue(),
        expected_tenant_id="tenant_acme",
        expected_month_id="month_acme_2026_08",
        year=2026,
        month=8,
        stores={"s1": "store_acme_s1"},
    )
    assert {"MERGED_CELL"} <= {e["code"] for e in parsed.errors}


def test_schema_mismatch_is_reported():
    data = _template()
    wb = load_workbook(BytesIO(data))
    wb["Manifest"]["B1"] = "OLD-SCHEMA"
    out = BytesIO()
    wb.save(out)
    parsed = parse_schedule(
        out.getvalue(),
        expected_tenant_id="tenant_acme",
        expected_month_id="month_acme_2026_08",
        year=2026,
        month=8,
        stores={"s1": "store_acme_s1"},
    )
    assert "SCHEMA_MISMATCH" in {e["code"] for e in parsed.errors}


def test_missing_token_is_parsed_as_empty():
    data = _template()
    wb = load_workbook(BytesIO(data))
    wb["Manifest"]["B5"] = None
    out = BytesIO()
    wb.save(out)
    parsed = parse_schedule(
        out.getvalue(),
        expected_tenant_id="tenant_acme",
        expected_month_id="month_acme_2026_08",
        year=2026,
        month=8,
        stores={"s1": "store_acme_s1"},
    )
    assert parsed.contract_token == ""
    # The API contract layer deterministically rejects an empty token
    # (covered by the service-level contract tests).
